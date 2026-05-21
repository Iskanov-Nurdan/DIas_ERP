from datetime import date
from io import BytesIO
from decimal import Decimal
import unittest

from django.contrib.auth import get_user_model
from rest_framework import status
from rest_framework.test import APITestCase

from apps.recipes.models import PlasticProfile
from apps.sales.models import Client, Order, OrderLine, Payment, Return, Sale, SaleLine
from apps.warehouse.models import WarehouseBatch


class SalesApiContractTests(APITestCase):
    def setUp(self):
        user_model = get_user_model()
        self.user = user_model.objects.create_superuser(
            email='sales-admin@example.com',
            password='pass12345',
            name='Sales Admin',
        )
        self.client.force_authenticate(self.user)
        self.client_active = Client.objects.create(name='Активный', is_active=True, credit_limit=Decimal('1000'))
        self.client_inactive = Client.objects.create(name='Неактивный', is_active=False)
        self.order = Order.objects.create(
            order_number='ORD-2026-001',
            date=date(2026, 4, 26),
            client=self.client_active,
            status=Order.STATUS_CONFIRMED,
        )
        self.order_line = OrderLine.objects.create(
            order=self.order,
            product='60 мм белый',
            ordered_quantity=Decimal('10'),
            shipped_quantity=Decimal('0'),
            unit_price=Decimal('100'),
        )
        self.good_batch = WarehouseBatch.objects.create(
            product='60 мм белый',
            quantity=Decimal('20'),
            date=date(2026, 4, 26),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_PACKED,
            pieces_per_package=Decimal('10'),
            packages_count=Decimal('2'),
        )
        self.defect_batch = WarehouseBatch.objects.create(
            product='Брак профиль',
            quantity=Decimal('10'),
            date=date(2026, 4, 26),
            quality=WarehouseBatch.QUALITY_DEFECT,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_UNPACKED,
        )

    def _payload(self, sale_status=Sale.STATUS_DRAFT, client_id=None, with_batch=True, qty='5', unit_price='100'):
        return {
            'date': '2026-04-26',
            'client': client_id if client_id is not None else self.client_active.pk,
            'linked_order': self.order.pk,
            'sale_status': sale_status,
            'comment': 'test',
            'sale_lines': [
                {
                    'order_line': self.order_line.pk,
                    'product': '60 мм белый',
                    'warehouse_batch': self.good_batch.pk if with_batch else None,
                    'stock_form': 'packed',
                    'piece_pick': 'from_sealed_package' if with_batch else '',
                    'quantity': qty,
                    'unit_price': unit_price,
                    'defect_flag': False,
                    'comment': '',
                }
            ],
        }

    def _create_sale(self, sale_status=Sale.STATUS_DRAFT):
        resp = self.client.post('/api/sales/', data=self._payload(sale_status=sale_status), format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        return Sale.objects.get(pk=resp.data['id'])

    def test_create_requires_client(self):
        payload = self._payload()
        payload.pop('client')
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'MISSING_CLIENT')

    def test_create_rejects_inactive_client(self):
        resp = self.client.post('/api/sales/', data=self._payload(client_id=self.client_inactive.pk), format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'INACTIVE_CLIENT')

    def test_create_requires_sale_lines(self):
        payload = self._payload()
        payload.pop('sale_lines')
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'MISSING_SALE_LINES')

    def test_create_rejects_header_only(self):
        payload = {'date': '2026-04-26', 'client': self.client_active.pk, 'product': 'X', 'quantity': '1', 'price': '1'}
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'MISSING_SALE_LINES')

    def test_create_rejects_quantity_non_positive(self):
        resp = self.client.post('/api/sales/', data=self._payload(qty='0'), format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'SALE_QUANTITY_INVALID')

    def test_create_rejects_negative_unit_price(self):
        resp = self.client.post('/api/sales/', data=self._payload(unit_price='-1'), format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'UNIT_PRICE_NEGATIVE')

    def test_create_rejects_closed_status(self):
        resp = self.client.post('/api/sales/', data=self._payload(sale_status=Sale.STATUS_CLOSED), format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'CLOSED_CREATE_FORBIDDEN')

    def test_create_draft_success(self):
        resp = self.client.post('/api/sales/', data=self._payload(sale_status=Sale.STATUS_DRAFT), format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)

    def test_create_draft_applies_warehouse_stock(self):
        before = Decimal(str(self.good_batch.quantity))
        resp = self.client.post('/api/sales/', data=self._payload(sale_status=Sale.STATUS_DRAFT, qty='5'), format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        sale = Sale.objects.get(pk=resp.data['id'])
        self.assertTrue(sale.warehouse_stock_applied)
        self.good_batch.refresh_from_db()
        self.assertLess(Decimal(str(self.good_batch.quantity)), before)

    def test_create_shipped_success_with_batch(self):
        resp = self.client.post('/api/sales/', data=self._payload(sale_status=Sale.STATUS_SHIPPED), format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        sale = Sale.objects.get(pk=resp.data['id'])
        self.assertTrue(sale.warehouse_stock_applied)

    def test_create_shipped_requires_batch_on_line(self):
        resp = self.client.post(
            '/api/sales/',
            data=self._payload(sale_status=Sale.STATUS_SHIPPED, with_batch=False),
            format='json',
        )
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'MISSING_WAREHOUSE_BATCH')

    def test_create_rejects_defect_batch_for_regular_sale(self):
        payload = self._payload(sale_status=Sale.STATUS_SHIPPED)
        payload['sale_lines'][0]['warehouse_batch'] = self.defect_batch.pk
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'DEFECT_BATCH_FORBIDDEN')

    def test_create_rejects_insufficient_stock(self):
        payload = self._payload(sale_status=Sale.STATUS_SHIPPED, qty='25')
        payload['sale_lines'][0]['order_line'] = None
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'INSUFFICIENT_STOCK')

    def test_create_rejects_order_line_exceeded(self):
        resp = self.client.post('/api/sales/', data=self._payload(qty='11'), format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'ORDER_LINE_QUANTITY_EXCEEDED')

    def test_create_with_order_full_shipping_sets_order_closed(self):
        payload = self._payload(qty='10')
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        self.order.refresh_from_db()
        self.order_line.refresh_from_db()
        self.assertEqual(self.order.status, Order.STATUS_CLOSED)
        self.assertEqual(self.order_line.shipped_quantity, Decimal('10'))

    def test_create_with_order_partial_shipping_sets_order_partially_shipped(self):
        payload = self._payload(qty='4')
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        self.order.refresh_from_db()
        self.order_line.refresh_from_db()
        self.assertEqual(self.order.status, Order.STATUS_PARTIALLY_SHIPPED)
        self.assertEqual(self.order_line.shipped_quantity, Decimal('4'))

    def test_create_with_order_line_autobind_when_product_text_differs(self):
        payload = self._payload(qty='3')
        payload['sale_lines'][0].pop('order_line', None)
        payload['sale_lines'][0]['product'] = ' 60 ММ БЕЛЫЙ '
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)

    def test_production_order_without_lines_closes_on_full_sale(self):
        prod_order = Order.objects.create(
            order_number='ORD-2026-PROD-NOLINES',
            date=date(2026, 4, 26),
            client=self.client_active,
            status=Order.STATUS_CONFIRMED,
            request_status=Order.REQUEST_STATUS_IN_PRODUCTION,
            production_quantity=4,
        )
        payload = self._payload(qty='4')
        payload['linked_order'] = prod_order.pk
        payload['sale_lines'][0].pop('order_line', None)
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        prod_order.refresh_from_db()
        self.assertEqual(prod_order.status, Order.STATUS_CLOSED)
        self.assertIsNone(prod_order.request_status)

    def test_update_rejects_status_via_patch(self):
        sale = self._create_sale()
        resp = self.client.patch(f'/api/sales/{sale.pk}/', data={'sale_status': Sale.STATUS_CONFIRMED}, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'SALE_STATUS_UPDATE_FORBIDDEN')

    def test_update_rejects_shipped_sale(self):
        sale = self._create_sale(sale_status=Sale.STATUS_SHIPPED)
        resp = self.client.patch(f'/api/sales/{sale.pk}/', data={'comment': 'x'}, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'SALE_UPDATE_FORBIDDEN')

    def test_update_rejects_when_active_payment(self):
        sale = self._create_sale()
        Payment.objects.create(
            client=self.client_active,
            linked_sale=sale,
            date=date(2026, 4, 26),
            payment_type=Payment.TYPE_PAYMENT,
            payment_method=Payment.METHOD_CASH,
            amount=Decimal('10'),
            status=Payment.STATUS_ACTIVE,
        )
        resp = self.client.patch(f'/api/sales/{sale.pk}/', data={'comment': 'x'}, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'SALE_LOCKED_BY_PAYMENT')

    def test_update_rejects_when_return_exists(self):
        sale = self._create_sale()
        Return.objects.create(sale=sale, linked_order=self.order, date=date(2026, 4, 26))
        resp = self.client.patch(f'/api/sales/{sale.pk}/', data={'comment': 'x'}, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'SALE_LOCKED_BY_RETURN')

    def test_update_safe_fields_allowed(self):
        sale = self._create_sale()
        resp = self.client.patch(
            f'/api/sales/{sale.pk}/',
            data={'comment': 'ok', 'invoice_number': 'INV-1', 'receipt_number': 'RCPT-1'},
            format='json',
        )
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_status_transition_and_missing_status(self):
        sale = self._create_sale()
        bad = self.client.patch(f'/api/sales/{sale.pk}/status/', data={}, format='json')
        self.assertEqual(bad.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(bad.data.get('code'), 'MISSING_STATUS')
        ok = self.client.patch(
            f'/api/sales/{sale.pk}/status/',
            data={'status': Sale.STATUS_CONFIRMED},
            format='json',
        )
        self.assertEqual(ok.status_code, status.HTTP_200_OK)

    def test_status_invalid_transition(self):
        sale = self._create_sale()
        resp = self.client.patch(
            f'/api/sales/{sale.pk}/status/',
            data={'status': Sale.STATUS_CLOSED},
            format='json',
        )
        self.assertEqual(resp.status_code, status.HTTP_422_UNPROCESSABLE_ENTITY)
        self.assertEqual(resp.data.get('code'), 'INVALID_STATUS_TRANSITION')

    def test_cancel_sale_guards_and_rollback(self):
        sale = self._create_sale(sale_status=Sale.STATUS_PARTIALLY_SHIPPED)
        Payment.objects.create(
            client=self.client_active,
            linked_sale=sale,
            date=date(2026, 4, 26),
            payment_type=Payment.TYPE_PAYMENT,
            payment_method=Payment.METHOD_CASH,
            amount=Decimal('10'),
            status=Payment.STATUS_ACTIVE,
        )
        blocked = self.client.post(f'/api/sales/{sale.pk}/cancel/', data={}, format='json')
        self.assertEqual(blocked.status_code, status.HTTP_409_CONFLICT)
        self.assertEqual(blocked.data.get('code'), 'HAS_PAYMENTS')
        Payment.objects.filter(linked_sale=sale).update(status=Payment.STATUS_CANCELED)
        ok = self.client.post(f'/api/sales/{sale.pk}/cancel/', data={}, format='json')
        self.assertEqual(ok.status_code, status.HTTP_200_OK)
        sale.refresh_from_db()
        self.assertEqual(sale.sale_status, Sale.STATUS_CANCELED)

    def test_select_sources_filters_and_shapes(self):
        other_client = Client.objects.create(name='Другой', is_active=True)
        other_order = Order.objects.create(
            order_number='ORD-2026-999',
            date=date(2026, 4, 26),
            client=other_client,
            status=Order.STATUS_CONFIRMED,
        )
        OrderLine.objects.create(order=other_order, product='X', ordered_quantity=Decimal('1'))
        resp = self.client.get(f'/api/sales/select-sources/?client_id={self.client_active.pk}&order_id={self.order.pk}')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertIn('order_lines', resp.data)
        self.assertTrue(all(row['id'] == self.order.pk for row in resp.data['orders']))
        self.assertTrue(all(b['quality'] == WarehouseBatch.QUALITY_GOOD for b in resp.data['warehouse_batches']))
        self.assertTrue(all('profile_id' in b for b in resp.data.get('available_warehouse_batches', [])))
        self.assertTrue(all('pieces_per_package' in b for b in resp.data.get('available_warehouse_batches', [])))
        self.assertTrue(all('client_id' in o for o in resp.data.get('orders', [])))
        order_row = resp.data['orders'][0]
        self.assertIn('paid_amount', order_row)
        self.assertIn('debt_amount', order_row)
        self.assertIn('payment_type', order_row)
        self.assertIn('payment_method', order_row)

    def test_select_sources_pieces_unpacked_without_packages_count_no_virtual_packages(self):
        wb = WarehouseBatch.objects.create(
            product='Профиль в куче',
            quantity=Decimal('38'),
            date=date(2026, 4, 26),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_UNPACKED,
            pieces_per_package=Decimal('10'),
            packages_count=None,
        )
        resp = self.client.get('/api/sales/select-sources/?unit_type=pieces')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        row = next((b for b in resp.data['available_warehouse_batches'] if b['id'] == wb.pk), None)
        self.assertIsNotNone(row)
        self.assertEqual(row['available_pieces'], '38')
        self.assertEqual(row['available_unpacked_pieces'], '38')
        self.assertEqual(row['unpacked_pieces'], '38')
        self.assertEqual(row['available_pieces_total'], '38')
        self.assertIsNone(row['available_packages'])

    def test_select_sources_pieces_mixed_open_package_splits_loose(self):
        wb = WarehouseBatch.objects.create(
            product='Смешанный остаток',
            quantity=Decimal('38'),
            date=date(2026, 4, 26),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_OPEN_PACKAGE,
            pieces_per_package=Decimal('10'),
            packages_count=Decimal('3'),
        )
        resp = self.client.get('/api/sales/select-sources/?unit_type=pieces')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        row = next(b for b in resp.data['available_warehouse_batches'] if b['id'] == wb.pk)
        self.assertEqual(row['available_pieces'], '8')
        self.assertEqual(row['available_unpacked_pieces'], '8')
        self.assertEqual(row['available_pieces_total'], '38')
        self.assertEqual(row['available_packages'], '3')

    def test_select_sources_pieces_all_sealed_open_package_excluded(self):
        wb = WarehouseBatch.objects.create(
            product='Только коробки',
            quantity=Decimal('30'),
            date=date(2026, 4, 26),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_OPEN_PACKAGE,
            pieces_per_package=Decimal('10'),
            packages_count=Decimal('3'),
        )
        resp = self.client.get('/api/sales/select-sources/?unit_type=pieces')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertNotIn(wb.pk, {b['id'] for b in resp.data['available_warehouse_batches']})

    def test_create_sale_applies_order_prepayment_without_double_payment_doc(self):
        Payment.objects.create(
            client=self.client_active,
            linked_order=self.order,
            date=date(2026, 4, 26),
            payment_type=Payment.TYPE_PREPAYMENT,
            payment_method=Payment.METHOD_CARD,
            amount=Decimal('300'),
            status=Payment.STATUS_ACTIVE,
        )
        payload = self._payload(qty='5', unit_price='100')
        payload.pop('payment_type', None)
        payload.pop('payment_method', None)
        payload.pop('paid_amount', None)
        payload['order_paid_amount_applied'] = '300'
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        sale = Sale.objects.get(pk=resp.data['id'])
        self.assertEqual(sale.order_paid_amount_applied, Decimal('300'))
        self.assertEqual(
            Payment.objects.filter(linked_sale=sale, status=Payment.STATUS_ACTIVE).count(),
            0,
        )

    def test_create_mixed_line_unit_types(self):
        payload = self._payload(qty='5', unit_price='100')
        payload['unit_type'] = 'pieces'
        payload['sale_lines'] = [
            {
                'warehouse_batch': self.good_batch.pk,
                'quantity': '0.5',
                'unit_price': '100',
                'unit_type': 'packages',
                'product': '60 мм белый',
            },
            {
                'warehouse_batch': self.good_batch.pk,
                'quantity': '3',
                'unit_price': '100',
                'unit_type': 'pieces',
                'product': '60 мм белый',
            },
        ]
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)

    def test_create_invalid_line_unit_type(self):
        payload = self._payload()
        payload['sale_lines'][0]['unit_type'] = 'meters'
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'INVALID_LINE_UNIT_TYPE')

    def test_preview_mixed_line_unit_types(self):
        payload = {
            'client': self.client_active.pk,
            'unit_type': 'pieces',
            'sale_lines': [
                {
                    'warehouse_batch': self.good_batch.pk,
                    'quantity': '1',
                    'unit_price': '100',
                    'unit_type': 'packages',
                },
                {
                    'warehouse_batch': self.good_batch.pk,
                    'quantity': '2',
                    'unit_price': '100',
                    'unit_type': 'pieces',
                },
            ],
            'payment_type': 'partial',
            'payment_method': 'cash',
            'paid_amount': '100',
        }
        resp = self.client.post('/api/sales/preview/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(len(resp.data.get('normalized_lines') or []), 2)

    def test_preview_without_payment_fields_defaults_to_debt(self):
        payload = {
            'client': self.client_active.pk,
            'unit_type': 'pieces',
            'sale_lines': [
                {
                    'warehouse_batch': self.good_batch.pk,
                    'quantity': '1',
                    'unit_price': '100',
                    'unit_type': 'pieces',
                },
            ],
        }
        resp = self.client.post('/api/sales/preview/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data.get('payment_status'), 'debt')

    def test_preview_accepts_client_id_alias(self):
        payload = {
            'client_id': self.client_active.pk,
            'unit_type': 'pieces',
            'sale_lines': [
                {
                    'warehouse_batch': self.good_batch.pk,
                    'quantity': '1',
                    'unit_price': '10',
                    'unit_type': 'pieces',
                },
            ],
        }
        resp = self.client.post('/api/sales/preview/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_create_sale_accepts_client_id_alias(self):
        payload = self._payload()
        cid = payload.pop('client')
        payload['client_id'] = cid
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)

    def test_create_sale_payment_method_without_payment_type(self):
        payload = self._payload()
        payload.pop('payment_type', None)
        payload['payment_method'] = 'transfer'
        resp = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)

    def test_waybill_and_receipt_are_html(self):
        sale = self._create_sale()
        w = self.client.get(f'/api/sales/{sale.pk}/waybill/')
        r = self.client.get(f'/api/sales/{sale.pk}/receipt/')
        self.assertEqual(w.status_code, status.HTTP_200_OK)
        self.assertEqual(r.status_code, status.HTTP_200_OK)
        self.assertIn('text/html', w['Content-Type'])
        self.assertIn('text/html', r['Content-Type'])

    def test_waybill_html_contains_required_sections(self):
        sale = self._create_sale()
        sale.refresh_from_db()
        resp = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=html')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        body = resp.content.decode('utf-8')
        self.assertIn('Расходная накладная №', body)
        self.assertIn('Поставщик:', body)
        self.assertIn('Покупатель:', body)
        self.assertIn('Итого', body)
        self.assertIn('от _________ г.', body)
        self.assertIn('тел: ____________________', body)

    def test_waybill_pdf_format(self):
        try:
            import xhtml2pdf  # noqa: F401
        except Exception:
            raise unittest.SkipTest('xhtml2pdf is not installed')
        sale = self._create_sale()
        resp = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=pdf')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertIn(f'waybill-{sale.pk}.pdf', resp['Content-Disposition'])

    def test_waybill_xlsx_format(self):
        sale = self._create_sale()
        resp = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=xlsx')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn(f'waybill-{sale.pk}.xlsx', resp['Content-Disposition'])

    def test_waybill_xlsx_layout_cells(self):
        from openpyxl import load_workbook

        sale = self._create_sale()
        resp = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=xlsx')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertIn(f'Расходная накладная № {sale.pk}', str(ws['A1'].value))
        self.assertEqual(ws['A3'].value, 'Поставщик: ____________________, тел: ____________________')
        self.assertIn('Покупатель:', str(ws['A4'].value))
        self.assertEqual(ws['A6'].value, '№')
        self.assertEqual(ws['B6'].value, 'Наименование товара')
        self.assertEqual(ws['C6'].value, 'Единица измерение')
        self.assertEqual(ws['D6'].value, 'Цена')
        self.assertEqual(ws['E6'].value, 'Сумма')
        self.assertEqual(ws['A11'].value, 'Отпустил')
        self.assertEqual(ws['C11'].value, 'Получил')
        self.assertEqual(ws['E11'].value, 'Место печати')
        self.assertEqual(ws['A13'].value, '____________________')
        self.assertEqual(ws['C13'].value, '____________________')
        self.assertEqual(ws['E13'].value, '____________________')

    def test_waybill_sale_not_found(self):
        resp = self.client.get('/api/sales/999999/waybill/?format=html')
        self.assertEqual(resp.status_code, status.HTTP_404_NOT_FOUND)
        self.assertEqual(resp.data.get('code'), 'SALE_NOT_FOUND')

    def test_waybill_invalid_format(self):
        sale = self._create_sale()
        resp = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=xml')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(resp.data.get('code'), 'INVALID_FORMAT')

    def test_waybill_json_format_strict_shape(self):
        sale = self._create_sale()
        resp = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data.get('title'), f'Расходная накладная № {sale.pk} от ________ г.')
        self.assertEqual(resp.data.get('buyer_name'), self.client_active.name)
        self.assertEqual(resp.data.get('date_line'), '________')
        self.assertEqual(resp.data.get('supplier_line'), '_______________________')
        self.assertEqual(resp.data.get('phone_line'), '_______________________')
        self.assertIsInstance(resp.data.get('sale_lines'), list)
        self.assertTrue(len(resp.data['sale_lines']) >= 1)
        row = resp.data['sale_lines'][0]
        self.assertIn('name', row)
        self.assertIn('quantity_display', row)
        self.assertIn('unit_price', row)
        self.assertIn('line_total', row)
        self.assertIn('total', resp.data)

    def test_waybill_json_by_accept_header(self):
        sale = self._create_sale()
        resp = self.client.get(
            f'/api/sales/{sale.pk}/waybill/',
            HTTP_ACCEPT='application/json',
        )
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.data.get('title'), f'Расходная накладная № {sale.pk} от ________ г.')

    def test_waybill_total_matches_sale_lines_sum(self):
        sale = self._create_sale()
        SaleLine.objects.filter(sale=sale).delete()
        SaleLine.objects.create(
            sale=sale,
            product='Товар 1',
            quantity=Decimal('2'),
            unit_price=Decimal('100'),
            line_total=Decimal('200'),
        )
        SaleLine.objects.create(
            sale=sale,
            product='Товар 2',
            quantity=Decimal('1'),
            unit_price=Decimal('150'),
            line_total=Decimal('150'),
        )
        sale.revenue = Decimal('350')
        sale.save(update_fields=['revenue'])
        resp = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=html')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        body = resp.content.decode('utf-8')
        self.assertIn('350', body)

    def test_waybill_packages_quantity_format_1x3_equals_3(self):
        batch = WarehouseBatch.objects.create(
            product='Пакет 3',
            quantity=Decimal('9'),
            date=date(2026, 4, 26),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_PACKED,
            pieces_per_package=Decimal('3'),
            packages_count=Decimal('3'),
        )
        sale = Sale.objects.create(
            order_number='WB-PACK-1',
            product='Пакет 3',
            sale_mode=Sale.MODE_PACKAGES,
            sold_packages=Decimal('1'),
            sold_pieces=Decimal('3'),
            quantity=Decimal('3'),
            date=date(2026, 4, 26),
            client=self.client_active,
            sale_status=Sale.STATUS_DRAFT,
            revenue=Decimal('300'),
        )
        SaleLine.objects.create(
            sale=sale,
            product='Пакет 3',
            warehouse_batch=batch,
            quantity=Decimal('3'),
            unit_price=Decimal('100'),
            line_total=Decimal('300'),
        )
        html = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=html')
        self.assertEqual(html.status_code, status.HTTP_200_OK)
        self.assertIn('1 упак × 3 шт = 3 шт', html.content.decode('utf-8'))
        from openpyxl import load_workbook
        wb_resp = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=xlsx')
        ws = load_workbook(BytesIO(wb_resp.content)).active
        self.assertEqual(ws['C7'].value, '1 упак × 3 шт = 3 шт')

    def test_waybill_packages_quantity_format_3x6_equals_18(self):
        batch = WarehouseBatch.objects.create(
            product='Пакет 6',
            quantity=Decimal('30'),
            date=date(2026, 4, 26),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_PACKED,
            pieces_per_package=Decimal('6'),
            packages_count=Decimal('5'),
        )
        sale = Sale.objects.create(
            order_number='WB-PACK-2',
            product='Пакет 6',
            sale_mode=Sale.MODE_PACKAGES,
            sold_packages=Decimal('3'),
            sold_pieces=Decimal('18'),
            quantity=Decimal('18'),
            date=date(2026, 4, 26),
            client=self.client_active,
            sale_status=Sale.STATUS_DRAFT,
            revenue=Decimal('1800'),
        )
        SaleLine.objects.create(
            sale=sale,
            product='Пакет 6',
            warehouse_batch=batch,
            quantity=Decimal('18'),
            unit_price=Decimal('100'),
            line_total=Decimal('1800'),
        )
        html = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=html')
        self.assertEqual(html.status_code, status.HTTP_200_OK)
        self.assertIn('3 упак × 6 шт = 18 шт', html.content.decode('utf-8'))

    def test_waybill_pieces_quantity_format_15(self):
        sale = Sale.objects.create(
            order_number='WB-PIECES-1',
            product='Штучный',
            sale_mode=Sale.MODE_PIECES,
            quantity=Decimal('15'),
            sold_pieces=Decimal('15'),
            date=date(2026, 4, 26),
            client=self.client_active,
            sale_status=Sale.STATUS_DRAFT,
            revenue=Decimal('1500'),
        )
        SaleLine.objects.create(
            sale=sale,
            product='Штучный',
            quantity=Decimal('15'),
            unit_price=Decimal('100'),
            line_total=Decimal('1500'),
        )
        html = self.client.get(f'/api/sales/{sale.pk}/waybill/?format=html')
        self.assertEqual(html.status_code, status.HTTP_200_OK)
        self.assertIn('15 шт', html.content.decode('utf-8'))

    def test_retrieve_sale_lines_include_packaging_fields_with_fallback(self):
        batch = WarehouseBatch.objects.create(
            product='Пакет fallback',
            quantity=Decimal('20'),
            date=date(2026, 4, 26),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_PACKED,
            pieces_per_package=Decimal('6'),
            packages_count=Decimal('5'),
        )
        sale = Sale.objects.create(
            order_number='SALE-RETRIEVE-PACK',
            product='Пакет fallback',
            sale_mode=Sale.MODE_PACKAGES,
            quantity=Decimal('12'),
            sold_pieces=Decimal('12'),
            date=date(2026, 4, 26),
            client=self.client_active,
            sale_status=Sale.STATUS_DRAFT,
            revenue=Decimal('1200'),
        )
        SaleLine.objects.create(
            sale=sale,
            product='Пакет fallback',
            warehouse_batch=batch,
            quantity=Decimal('12'),
            unit_price=Decimal('100'),
            line_total=Decimal('1200'),
        )
        resp = self.client.get(f'/api/sales/{sale.pk}/')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        line = resp.data['sale_lines'][0]
        self.assertEqual(line.get('unit_type'), 'packages')
        self.assertEqual(line.get('quantity'), '12')
        self.assertEqual(line.get('packages_quantity'), '2')
        self.assertEqual(line.get('pieces_per_package'), '6')
        self.assertEqual(line.get('warehouse_batch'), batch.pk)
        self.assertIn('Пакет fallback', line.get('warehouse_batch_display') or '')
        self.assertEqual(line.get('quantity_display'), '2 упак × 6 шт = 12 шт')
        self.assertIn('order_line', line)
        self.assertIn('order_line_display', line)

    def test_multi_profile_order_select_sources_and_sale_two_lines(self):
        profile_a = PlasticProfile.objects.create(name='Профиль 5м', code='P5', is_active=True)
        profile_b = PlasticProfile.objects.create(name='Профиль 6м', code='P6', is_active=True)
        order = Order.objects.create(
            order_number='ORD-MULTI-01',
            date=date(2026, 5, 21),
            client=self.client_active,
            status=Order.STATUS_CONFIRMED,
        )
        line_a = OrderLine.objects.create(
            order=order,
            product=profile_a.name,
            profile=profile_a,
            ordered_quantity=Decimal('20'),
            unit_price=Decimal('150'),
        )
        line_b = OrderLine.objects.create(
            order=order,
            product=profile_b.name,
            profile=profile_b,
            ordered_quantity=Decimal('10'),
            unit_price=Decimal('180'),
        )
        batch_a = WarehouseBatch.objects.create(
            product=profile_a.name,
            profile=profile_a,
            quantity=Decimal('50'),
            date=date(2026, 5, 21),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_UNPACKED,
        )
        batch_b = WarehouseBatch.objects.create(
            product=profile_b.name,
            profile=profile_b,
            quantity=Decimal('30'),
            date=date(2026, 5, 21),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_UNPACKED,
        )
        sources = self.client.get(f'/api/sales/select-sources/?client={self.client_active.pk}')
        self.assertEqual(sources.status_code, status.HTTP_200_OK)
        avail = sources.data.get('available_orders') or []
        row = next(x for x in avail if x['id'] == order.pk)
        self.assertEqual(row.get('lines_count'), 2)
        self.assertEqual(len(row.get('order_lines') or []), 2)
        self.assertIn('2026-05-21', row.get('display', ''))
        self.assertIn('2', row.get('display', ''))
        self.assertEqual(len(avail), len({x['id'] for x in avail}))

        detail = self.client.get(f'/api/orders/{order.pk}/')
        self.assertEqual(len(detail.data.get('order_lines') or []), 2)

        payload = {
            'date': '2026-05-21',
            'client': self.client_active.pk,
            'order': order.pk,
            'sale_status': Sale.STATUS_DRAFT,
            'sale_lines': [
                {
                    'order_line': line_a.pk,
                    'warehouse_batch': batch_a.pk,
                    'quantity': '20',
                    'unit_price': '150',
                    'unit_type': 'pieces',
                    'product': profile_a.name,
                },
                {
                    'order_line': line_b.pk,
                    'warehouse_batch': batch_b.pk,
                    'quantity': '10',
                    'unit_price': '180',
                    'unit_type': 'pieces',
                    'product': profile_b.name,
                },
            ],
        }
        preview = self.client.post('/api/sales/preview/', data=payload, format='json')
        self.assertEqual(preview.status_code, status.HTTP_200_OK)
        create = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(create.status_code, status.HTTP_201_CREATED)
        sale = Sale.objects.get(pk=create.data['id'])
        slines = list(sale.sale_lines.order_by('order_line_id'))
        self.assertEqual(len(slines), 2)
        self.assertEqual({sl.order_line_id for sl in slines}, {line_a.pk, line_b.pk})

    def test_order_partial_prepayment_sale_supplemental_no_double_revenue(self):
        profile_a = PlasticProfile.objects.create(name='Профиль A', code='PA', is_active=True)
        profile_b = PlasticProfile.objects.create(name='Профиль B', code='PB', is_active=True)
        order = Order.objects.create(
            order_number='ORD-PAY-01',
            date=date(2026, 5, 21),
            client=self.client_active,
            status=Order.STATUS_CONFIRMED,
        )
        line_a = OrderLine.objects.create(
            order=order,
            product=profile_a.name,
            profile=profile_a,
            ordered_quantity=Decimal('20'),
            unit_price=Decimal('4000'),
        )
        line_b = OrderLine.objects.create(
            order=order,
            product=profile_b.name,
            profile=profile_b,
            ordered_quantity=Decimal('10'),
            unit_price=Decimal('2000'),
        )
        Payment.objects.create(
            client=self.client_active,
            linked_order=order,
            date=date(2026, 5, 21),
            payment_type=Payment.TYPE_PREPAYMENT,
            payment_method=Payment.METHOD_CASH,
            amount=Decimal('40000'),
            status=Payment.STATUS_ACTIVE,
        )
        batch_a = WarehouseBatch.objects.create(
            product=profile_a.name,
            profile=profile_a,
            quantity=Decimal('50'),
            date=date(2026, 5, 21),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_UNPACKED,
        )
        batch_b = WarehouseBatch.objects.create(
            product=profile_b.name,
            profile=profile_b,
            quantity=Decimal('30'),
            date=date(2026, 5, 21),
            quality=WarehouseBatch.QUALITY_GOOD,
            status=WarehouseBatch.STATUS_AVAILABLE,
            inventory_form=WarehouseBatch.INVENTORY_UNPACKED,
        )
        sources = self.client.get(f'/api/sales/select-sources/?client={self.client_active.pk}')
        self.assertEqual(sources.status_code, status.HTTP_200_OK)
        row = next(x for x in sources.data['available_orders'] if x['id'] == order.pk)
        self.assertEqual(Decimal(row['paid_amount']), Decimal('40000'))
        self.assertEqual(Decimal(row['amount_remaining']), Decimal('60000'))
        self.assertEqual(Decimal(row['total_amount']), Decimal('100000'))
        self.assertEqual(row['payment_type'], 'partial')
        self.assertEqual(len(row['order_lines']), 2)

        detail = self.client.get(f'/api/orders/{order.pk}/')
        self.assertEqual(Decimal(detail.data['amount_remaining']), Decimal('60000'))
        self.assertEqual(Decimal(detail.data['paid_amount']), Decimal('40000'))

        payload = {
            'date': '2026-05-21',
            'client': self.client_active.pk,
            'order': order.pk,
            'payment_type': 'partial',
            'payment_method': 'cash',
            'paid_amount': '60000',
            'order_paid_amount_applied': '40000',
            'sale_lines': [
                {
                    'order_line': line_a.pk,
                    'warehouse_batch': batch_a.pk,
                    'quantity': '20',
                    'unit_price': '4000',
                    'unit_type': 'pieces',
                    'product': profile_a.name,
                },
                {
                    'order_line': line_b.pk,
                    'warehouse_batch': batch_b.pk,
                    'quantity': '10',
                    'unit_price': '2000',
                    'unit_type': 'pieces',
                    'product': profile_b.name,
                },
            ],
        }
        preview = self.client.post('/api/sales/preview/', data=payload, format='json')
        self.assertEqual(preview.status_code, status.HTTP_200_OK)
        self.assertEqual(Decimal(preview.data['paid_amount']), Decimal('60000'))
        self.assertEqual(Decimal(preview.data['order_paid_amount_applied']), Decimal('40000'))
        self.assertEqual(Decimal(preview.data['amount_remaining']), Decimal('0'))

        create = self.client.post('/api/sales/', data=payload, format='json')
        self.assertEqual(create.status_code, status.HTTP_201_CREATED)
        sale = Sale.objects.get(pk=create.data['id'])
        self.assertEqual(sale.order_paid_amount_applied, Decimal('40000'))
        self.assertEqual(sale.revenue, Decimal('100000'))
        sale_pay = Payment.objects.filter(
            linked_sale=sale,
            status=Payment.STATUS_ACTIVE,
            payment_type=Payment.TYPE_PAYMENT,
        )
        self.assertEqual(sale_pay.count(), 1)
        self.assertEqual(sale_pay.first().amount, Decimal('60000'))
        order_prepay = Payment.objects.filter(
            linked_order=order,
            status=Payment.STATUS_ACTIVE,
            payment_type=Payment.TYPE_PREPAYMENT,
        )
        self.assertEqual(order_prepay.count(), 1)
        self.assertEqual(order_prepay.first().amount, Decimal('40000'))

        from apps.sales.payment_status import sale_payment_metrics

        metrics = sale_payment_metrics(sale)
        self.assertEqual(metrics['debt_amount'], Decimal('0'))
        self.assertEqual(metrics['paid_amount'], Decimal('100000'))

    def test_credit_check_hard_block_and_override_access(self):
        self.client_active.credit_limit_mode = 'hard'
        self.client_active.credit_limit = Decimal('1')
        self.client_active.save(update_fields=['credit_limit_mode', 'credit_limit'])
        sale = self._create_sale(sale_status=Sale.STATUS_CONFIRMED)
        resp = self.client.patch(
            f'/api/sales/{sale.pk}/status/',
            data={'status': Sale.STATUS_SHIPPED, 'force_credit_override': True},
            format='json',
        )
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertTrue(resp.data.get('credit_limit_bypassed'))

