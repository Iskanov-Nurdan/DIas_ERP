import logging
from io import BytesIO
from pathlib import Path
from decimal import Decimal, InvalidOperation
from html import escape

from django.conf import settings
from django.db import transaction
from django.db.models import Count, DecimalField, Q, Sum, Value
from django.db.models.functions import Coalesce
from django.http import Http404, HttpResponse
from django.utils import timezone

from rest_framework import viewsets, status
from rest_framework.exceptions import ValidationError
from rest_framework.decorators import action
from rest_framework.response import Response
from drf_spectacular.utils import extend_schema, extend_schema_view, inline_serializer, OpenApiParameter
from rest_framework import serializers as drf_serializers

from apps.activity.mixins import ActivityLoggingMixin
from config.api_numbers import api_decimal_str
from config.pagination import WarehouseResultsSetPagination
from config.permissions import IsAdminOrHasAccess
from .filters import ClientFilter, DefectFilter, OrderFilter, PaymentFilter, ReturnFilter, SaleFilter
from .models import (
    Client,
    ClientPrice,
    DefectRecord,
    Order,
    OrderLine,
    OrderReservation,
    Payment,
    PriceList,
    Return,
    ReturnLine,
    ReworkRequest,
    Sale,
    SaleLine,
    Shipment,
)
from .client_order_production import apply_resource_check_to_order
from .serializers import (
    ClientPriceSerializer,
    ClientSerializer,
    DefectRecordSerializer,
    OrderReservationSerializer,
    OrderSerializer,
    OrderLineSerializer,
    PaymentSerializer,
    PriceListSerializer,
    ReturnSerializer,
    ReworkRequestSerializer,
    SaleSerializer,
    defect_record_source_label,
    rework_quantities_from_defect_record,
    return_document_amount,
    sale_active_total_amount,
    sale_return_source_is_valid,
    _reserve_defect_from_kg_only,
)

logger = logging.getLogger(__name__)


def _err(code: str, message: str, errors: list = None, http_status: int = 400) -> Response:
    """Единый стиль: строковые code / error / detail (для UI), опционально errors."""
    payload = {
        'code': code,
        'message': message,
        'error': message,
        'detail': message,
    }
    if errors:
        payload['fields'] = errors
        payload['errors'] = errors
    return Response(payload, status=http_status)


def _waybill_supplier_payload() -> dict:
    return {
        'name': '____________________',
        'phone': '____________________',
    }


def _waybill_line_name(line: SaleLine) -> str:
    warehouse_batch_display = ''
    if line.warehouse_batch_id and getattr(line.warehouse_batch, 'profile_id', None):
        profile_name = (line.warehouse_batch.profile.name or '').strip()
        if profile_name:
            warehouse_batch_display = profile_name
    return (
        warehouse_batch_display
        or (line.product or '').strip()
        or '—'
    )


def _safe_decimal(value):
    if value in (None, ''):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _first_decimal_attr(obj, names):
    for name in names:
        if not hasattr(obj, name):
            continue
        val = _safe_decimal(getattr(obj, name))
        if val is not None:
            return val
    return None


def _is_packages_mode(sale: Sale, line: SaleLine) -> bool:
    package_aliases = {'packages', 'упаковки', 'упак'}
    line_mode = str(getattr(line, 'unit_type', '') or '').strip().lower()
    sale_mode = str(getattr(sale, 'unit_type', '') or '').strip().lower()
    if line_mode in package_aliases or sale_mode in package_aliases:
        return True
    return sale.sale_mode == Sale.MODE_PACKAGES


def _warehouse_batch_select_sources_availability(b, raw_available: Decimal) -> dict:
    """
    Разбивка остатка для GET sales/select-sources/.

    - total_pieces: всегда get_available_quantity (штуки).
    - unpacked_pieces: неупакованный хвост для вкладки «штуки».
    - available_packages_str: только если это не приведёт к двойному вычитанию на фронте:
      для packed — raw / pieces_per_package;
      для unpacked/open_package — только при известном packages_count > 0 (запечатанные коробки),
      иначе None (даже при pieces_per_package > 0).
    """
    from apps.warehouse.models import WarehouseBatch

    raw = Decimal(str(raw_available))
    ppp = (
        Decimal(str(b.pieces_per_package))
        if b.pieces_per_package is not None
        else Decimal('0')
    )
    is_packed = b.inventory_form == WarehouseBatch.INVENTORY_PACKED

    if is_packed:
        unpacked = Decimal('0')
        if ppp > 0:
            pkg = (raw / ppp).quantize(Decimal('0.0001'))
            avail_pkg = api_decimal_str(pkg) if pkg > 0 else None
        else:
            avail_pkg = None
        return {'total_pieces': raw, 'unpacked_pieces': unpacked, 'available_packages_str': avail_pkg}

    pc_raw = b.packages_count
    if ppp > 0 and pc_raw is not None and Decimal(str(pc_raw)) > 0:
        pc = Decimal(str(pc_raw))
        sealed_pieces = min(raw, (pc * ppp)).quantize(Decimal('0.0001'))
        unpacked = (raw - sealed_pieces).quantize(Decimal('0.0001'))
        if unpacked < 0:
            unpacked = Decimal('0')
        if sealed_pieces > 0 and ppp > 0:
            sealed_pkg = (sealed_pieces / ppp).quantize(Decimal('0.0001'))
            avail_pkg = api_decimal_str(sealed_pkg) if sealed_pkg > 0 else None
        else:
            avail_pkg = None
    else:
        unpacked = raw
        avail_pkg = None

    return {'total_pieces': raw, 'unpacked_pieces': unpacked, 'available_packages_str': avail_pkg}


def _sale_line_packaging_payload(sale: Sale, line: SaleLine, lines_count: int) -> dict:
    pieces_qty = _first_decimal_attr(line, ('quantity', 'qty', 'pieces_quantity', 'quantity_pieces'))
    packages_qty = _first_decimal_attr(line, ('packages_quantity', 'quantity_packages', 'packages'))
    if packages_qty is None and lines_count == 1:
        sold_packages = _safe_decimal(getattr(sale, 'sold_packages', None))
        if sold_packages is not None and sold_packages > 0:
            packages_qty = sold_packages
    ppp = _first_decimal_attr(line, ('pieces_per_package', 'pieces_per_pack', 'pack_size', 'package_size', 'pack_qty'))
    if ppp is None and line.warehouse_batch_id:
        ppp = _first_decimal_attr(
            line.warehouse_batch,
            ('pieces_per_package', 'pieces_per_pack', 'pack_size', 'package_size', 'pack_qty'),
        )
    if packages_qty is None and ppp is not None and ppp > 0 and pieces_qty is not None and pieces_qty > 0:
        packages_qty = (pieces_qty / ppp)
    if ppp is None and packages_qty and packages_qty > 0 and pieces_qty and pieces_qty > 0:
        ppp = (pieces_qty / packages_qty)

    is_packages = _is_packages_mode(sale, line)
    unit_type = 'packages' if is_packages else 'pieces'
    quantity_display = '0 шт'
    if is_packages:
        if packages_qty is not None and ppp is not None and pieces_qty is not None:
            quantity_display = (
                f"{api_decimal_str(packages_qty)} упак × {api_decimal_str(ppp)} шт = {api_decimal_str(pieces_qty)} шт"
            )
        elif packages_qty is not None and pieces_qty is not None:
            quantity_display = f"{api_decimal_str(packages_qty)} упак = {api_decimal_str(pieces_qty)} шт"
        elif packages_qty is not None:
            quantity_display = f"{api_decimal_str(packages_qty)} упак"
        elif pieces_qty is not None:
            quantity_display = f"{api_decimal_str(pieces_qty)} шт"
    elif pieces_qty is not None:
        quantity_display = f"{api_decimal_str(pieces_qty)} шт"

    return {
        'unit_type': unit_type,
        'quantity': api_decimal_str(pieces_qty or Decimal('0')),
        'packages_quantity': (api_decimal_str(packages_qty) if packages_qty is not None else None),
        'pieces_per_package': (api_decimal_str(ppp) if ppp is not None else None),
        'quantity_display': quantity_display,
    }


def _format_waybill_quantity(sale: Sale, line: SaleLine, lines_count: int) -> str:
    return _sale_line_packaging_payload(sale, line, lines_count=lines_count)['quantity_display']


def _ensure_sale_waybill_number(sale: Sale) -> str:
    if sale.invoice_number:
        return sale.invoice_number
    doc_date = sale.date or (sale.created_at.date() if sale.created_at else None)
    year = doc_date.year if doc_date else 0
    generated = f'WB-{year}-{sale.id:06d}'
    Sale.objects.filter(pk=sale.pk, invoice_number='').update(invoice_number=generated)
    sale.invoice_number = generated
    return generated


def _build_sale_waybill_payload(sale: Sale) -> dict:
    _ensure_sale_waybill_number(sale)

    unit_label = 'шт'
    if sale.sale_mode == Sale.MODE_PACKAGES:
        unit_label = 'упак'

    lines_payload = []
    lines_total = Decimal('0')
    sale_lines = list(sale.sale_lines.select_related('warehouse_batch__profile').all())
    lines_count = len(sale_lines)
    for idx, line in enumerate(sale_lines, 1):
        price = Decimal(str(line.unit_price or 0)).quantize(Decimal('0.01'))
        row_total = Decimal(str(line.line_total or 0)).quantize(Decimal('0.01'))
        lines_total += row_total
        lines_payload.append({
            'index': idx,
            'name': _waybill_line_name(line),
            'quantity_display': _format_waybill_quantity(sale, line, lines_count=lines_count),
            'unit': _format_waybill_quantity(sale, line, lines_count=lines_count),
            'price': api_decimal_str(price),
            'sum': api_decimal_str(row_total),
        })

    if not lines_payload:
        legacy_price = Decimal(str(sale.price or 0)).quantize(Decimal('0.01'))
        legacy_sum = (legacy_price * Decimal(str(sale.quantity or 0))).quantize(Decimal('0.01'))
        lines_total = legacy_sum
        lines_payload.append({
            'index': 1,
            'name': (sale.product or '').strip() or '—',
            'quantity_display': f"{api_decimal_str(Decimal(str(sale.quantity or 0)))} {'упак' if sale.sale_mode == Sale.MODE_PACKAGES else 'шт'}",
            'unit': f"{api_decimal_str(Decimal(str(sale.quantity or 0)))} {'упак' if sale.sale_mode == Sale.MODE_PACKAGES else 'шт'}",
            'price': api_decimal_str(legacy_price),
            'sum': api_decimal_str(legacy_sum),
        })

    sale_total = Decimal(str(sale.revenue or 0)).quantize(Decimal('0.01'))
    total = lines_total if lines_total > 0 else sale_total
    client_name = sale.client.name if sale.client_id else '—'
    return {
        'sale_id': sale.id,
        'waybill_number': str(sale.id),
        'waybill_date': '_________',
        'supplier': _waybill_supplier_payload(),
        'buyer_name': client_name,
        'lines': lines_payload,
        'total': api_decimal_str(total),
    }


def _waybill_json_payload(payload: dict) -> dict:
    return {
        'title': payload.get('title') or f"Расходная накладная № {payload['sale_id']} от ________ г.",
        'buyer_name': payload.get('buyer_name') or '—',
        'date_line': '________',
        'supplier_line': '_______________________',
        'phone_line': '_______________________',
        'sale_lines': [
            {
                'name': str(line.get('name') or '—'),
                'quantity_display': str(line.get('quantity_display') or line.get('unit') or '0 шт'),
                'unit_price': str(line.get('price') or '0'),
                'line_total': str(line.get('sum') or '0'),
            }
            for line in payload.get('lines') or []
        ],
        'total': str(payload.get('total') or '0'),
    }


def _waybill_layout_spec(payload: dict) -> dict:
    return {
        'title': f"Расходная накладная № {payload['waybill_number']} от {payload['waybill_date']} г.",
        'supplier': f"Поставщик: {payload['supplier']['name']}, тел: {payload['supplier']['phone']}",
        'buyer': f"Покупатель: {payload['buyer_name']}",
        'headers': ['№', 'Наименование товара', 'Единица измерение', 'Цена', 'Сумма'],
        'signatures': [
            'Отпустил',
            'Получил',
            'Место печати',
        ],
        'signature_line': '____________________',
        'total_label': 'Итого',
    }


def _sale_waybill_pdf_response(payload: dict) -> HttpResponse:
    from xhtml2pdf import pisa

    html = _render_waybill_html(payload)
    pdf_buffer = BytesIO()
    result = pisa.CreatePDF(
        src=html,
        dest=pdf_buffer,
        encoding='utf-8',
    )
    if result.err:
        raise RuntimeError('PDF_RENDER_FAILED')
    pdf_buffer.seek(0)
    response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="waybill-{payload["sale_id"]}.pdf"'
    return response


def _sale_waybill_xlsx_response(payload: dict) -> HttpResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    layout = _waybill_layout_spec(payload)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Waybill'

    ws.merge_cells('A1:E1')
    ws['A1'] = layout['title']
    ws['A1'].font = Font(name='Times New Roman', bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A3:E3')
    ws['A3'] = layout['supplier']
    ws.merge_cells('A4:E4')
    ws['A4'] = layout['buyer']
    ws['A3'].font = Font(name='Times New Roman', size=12)
    ws['A4'].font = Font(name='Times New Roman', size=12)

    headers = layout['headers']
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(name='Times New Roman', bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    for line in payload['lines']:
        row += 1
        row_cells = [
            line['index'],
            line['name'],
            line['unit'],
            line['price'],
            line['sum'],
        ]
        for col, value in enumerate(row_cells, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = Font(name='Times New Roman', size=12)
            cell.border = border
            if col in (1, 3):
                cell.alignment = Alignment(horizontal='center')
            if col in (4, 5):
                cell.alignment = Alignment(horizontal='right')

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    total_label_cell = ws.cell(row=row, column=1, value=layout['total_label'])
    total_label_cell.font = Font(name='Times New Roman', bold=True, size=12)
    total_label_cell.alignment = Alignment(horizontal='right')
    total_label_cell.border = border
    for col in (2, 3, 4):
        ws.cell(row=row, column=col).border = border
    total_value_cell = ws.cell(row=row, column=5, value=payload['total'])
    total_value_cell.font = Font(name='Times New Roman', bold=True, size=12)
    total_value_cell.alignment = Alignment(horizontal='right')
    total_value_cell.border = border

    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value=layout['signatures'][0])
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3, value=layout['signatures'][1])
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=5)
    ws.cell(row=row, column=5, value=layout['signatures'][2])
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value=layout['signature_line'])
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3, value=layout['signature_line'])
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=5)
    ws.cell(row=row, column=5, value=layout['signature_line'])
    for coord in ('A11', 'C11', 'E11', 'A13', 'C13', 'E13'):
        ws[coord].font = Font(name='Times New Roman', size=12)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 18

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="waybill-{payload["sale_id"]}.xlsx"'
    return response


def _render_waybill_html(payload: dict) -> str:
    template_path = Path(settings.BASE_DIR) / 'apps' / 'sales' / 'templates' / 'sales' / 'waybill_template.html'
    template_raw = template_path.read_text(encoding='utf-8')
    layout = _waybill_layout_spec(payload)
    rows = []
    for line in payload['lines']:
        rows.append(
            '<tr>'
            f'<td class="num">{line["index"]}</td>'
            f'<td>{escape(str(line["name"]))}</td>'
            f'<td>{escape(str(line["unit"]))}</td>'
            f'<td class="money">{escape(str(line["price"]))}</td>'
            f'<td class="money">{escape(str(line["sum"]))}</td>'
            '</tr>'
        )
    lines_rows = ''.join(rows)
    substitutions = {
        'title': escape(layout['title']),
        'supplier': escape(layout['supplier']),
        'buyer': escape(layout['buyer']),
        'h1': escape(layout['headers'][0]),
        'h2': escape(layout['headers'][1]),
        'h3': escape(layout['headers'][2]),
        'h4': escape(layout['headers'][3]),
        'h5': escape(layout['headers'][4]),
        'total_label': escape(layout['total_label']),
        'sign1': escape(layout['signatures'][0]),
        'sign2': escape(layout['signatures'][1]),
        'sign3': escape(layout['signatures'][2]),
        'sign_line': escape(layout['signature_line']),
        'lines_rows': lines_rows,
        'total': escape(str(payload['total'])),
    }
    html = template_raw
    for key, value in substitutions.items():
        html = html.replace(f'{{{key}}}', value)
    return html


def _order_display_payload(order: Order) -> dict:
    from .payment_status import order_payment_metrics

    profile_name = None
    length = None
    qty = None
    total_meters = None
    request_status = getattr(order, 'request_status', None)
    if getattr(order, 'production_profile_id', None):
        profile_name = (order.production_profile.name if order.production_profile_id else None) or None
    if getattr(order, 'production_length', None) is not None:
        length = api_decimal_str(Decimal(str(order.production_length)))
    if getattr(order, 'production_quantity', None) is not None:
        qty = int(order.production_quantity)
    if getattr(order, 'request_total_meters', None) is not None:
        total_meters = api_decimal_str(Decimal(str(order.request_total_meters)))
    elif length is not None and qty is not None:
        total_meters = api_decimal_str((Decimal(str(length)) * Decimal(qty)).quantize(Decimal('0.0001')))
    shipping_statuses = {
        Order.STATUS_PARTIALLY_SHIPPED,
        Order.STATUS_SHIPPED,
        Order.STATUS_CLOSED,
    }
    prefer_order_status = order.status in shipping_statuses
    status_label = (
        order.get_status_display()
        if (prefer_order_status or not request_status)
        else order.get_request_status_display()
    )
    display_parts = [profile_name or '—']
    if qty is not None and length is not None:
        display_parts.append(f'{qty} шт × {length} м')
    elif qty is not None:
        display_parts.append(f'{qty} шт')
    if request_status and not prefer_order_status:
        display_parts.append(status_label.lower())
    payment_metrics = order_payment_metrics(order)
    paid_amount = Decimal(str(payment_metrics['paid_amount'] or 0)).quantize(Decimal('0.01'))
    debt_amount = Decimal(str(payment_metrics['debt_amount'] or 0)).quantize(Decimal('0.01'))
    total_amount = Decimal(str(order.total_amount or 0)).quantize(Decimal('0.01'))
    if debt_amount == 0:
        payment_type = 'full'
    elif paid_amount > 0:
        payment_type = 'partial'
    else:
        payment_type = 'debt'
    latest = (
        order.payments.filter(
            status=Payment.STATUS_ACTIVE,
            payment_type__in=(Payment.TYPE_PREPAYMENT, Payment.TYPE_PAYMENT, Payment.TYPE_SURCHARGE),
        )
        .order_by('-id')
        .first()
    )
    payment_method = None
    if latest is not None:
        payment_method = (
            'cash' if latest.payment_method == Payment.METHOD_CASH
            else ('card' if latest.payment_method == Payment.METHOD_CARD
            else ('transfer' if latest.payment_method == Payment.METHOD_TRANSFER else None))
        )
    recipe_name = None
    if order.resolved_recipe_id:
        recipe_name = (order.resolved_recipe.recipe or '').strip() or (order.resolved_recipe.product or '')[:255]
    return {
        'id': order.id,
        'display': ' — '.join(display_parts),
        'order_number': order.order_number,
        'date': order.date.isoformat() if order.date else None,
        'order_display': ' — '.join(display_parts),
        'client_name': order.client.name if order.client_id else None,
        'profile_name': profile_name,
        'quantity': qty,
        'length': length,
        'total_meters': total_meters,
        'request_status': request_status,
        'status_label': status_label,
        'recipe_id': order.resolved_recipe_id,
        'recipe_name': recipe_name,
        'paid_amount': api_decimal_str(paid_amount),
        'debt_amount': api_decimal_str(debt_amount),
        'total_amount': api_decimal_str(total_amount),
        'payment_type': payment_type,
        'payment_method': payment_method,
        'prepayment_amount': api_decimal_str(paid_amount),
        'advance_amount': api_decimal_str(paid_amount),
    }


# ─────────────────────────────────────────────────────────────────────────────
# CLIENT
# ─────────────────────────────────────────────────────────────────────────────

class ClientViewSet(ActivityLoggingMixin, viewsets.ModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    pagination_class = WarehouseResultsSetPagination
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'clients'
    activity_section = 'Клиенты'
    activity_label = 'клиент'
    filterset_class = ClientFilter
    search_fields = ['name', 'inn', 'settlement_account', 'contact', 'email', 'messenger']
    ordering_fields = ['id', 'name']

    def get_queryset(self):
        return (
            Client.objects.annotate(
                sales_count=Count(
                    'sales',
                    distinct=False,
                    filter=~Q(sales__sale_status__in=(Sale.STATUS_DRAFT, Sale.STATUS_CANCELED)),
                ),
                sales_total=Coalesce(
                    Sum(
                        'sales__revenue',
                        filter=~Q(sales__sale_status__in=(Sale.STATUS_DRAFT, Sale.STATUS_CANCELED)),
                    ),
                    Value(Decimal('0')),
                    output_field=DecimalField(max_digits=20, decimal_places=2),
                ),
            )
        )

    def destroy(self, request, *args, **kwargs):
        return Response(
            {
                'code': 'DELETE_DISABLED',
                'error': 'Физическое удаление клиентов отключено. Используйте is_active=false.',
                'detail': 'Патч клиента: {"is_active": false}.',
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    @action(detail=True, methods=['get'], url_path='history')
    def history(self, request, pk=None):
        """
        Агрегированная история клиента: заявки, продажи, оплаты, возвраты,
        долги, авансы, кредитный лимит, прибыль, просроченные долги.
        """
        from django.db.models import Sum
        from django.db.models.functions import Coalesce
        from config.api_numbers import api_decimal_str
        from .credit_check import compute_client_debt, check_credit_limit

        client = self.get_object()

        orders = Order.objects.filter(client=client).prefetch_related('lines', 'payments').order_by('-date')
        sales = Sale.objects.filter(client=client).order_by('-date')
        sales_for_aggregates = sales.exclude(
            sale_status__in=(Sale.STATUS_DRAFT, Sale.STATUS_CANCELED),
        )
        payments = Payment.objects.filter(client=client, status=Payment.STATUS_ACTIVE).order_by('-date')
        returns = Return.objects.filter(sale__client=client).order_by('-date')

        # Денежная аналитика
        total_paid = sum(
            (p.amount or Decimal('0'))
            for p in payments
            if p.payment_type in (Payment.TYPE_PREPAYMENT, Payment.TYPE_PAYMENT, Payment.TYPE_SURCHARGE)
        )
        total_refunded = sum(
            (p.amount or Decimal('0'))
            for p in payments
            if p.payment_type == Payment.TYPE_REFUND
        )
        net_paid = total_paid - total_refunded
        total_revenue = sum((s.revenue or Decimal('0')) for s in sales_for_aggregates)

        # Долг или аванс
        if net_paid < total_revenue:
            client_debt_money = total_revenue - net_paid
            client_advance_amount = Decimal('0')
        else:
            client_debt_money = Decimal('0')
            client_advance_amount = net_paid - total_revenue

        # Прибыль по клиенту
        total_cost = sum((s.cost or Decimal('0')) for s in sales_for_aggregates if not s.is_defect_sale)
        defect_revenue = sum((s.revenue or Decimal('0')) for s in sales_for_aggregates if s.is_defect_sale)
        total_profit = total_revenue - total_cost

        # Неотгруженные товары
        has_unshipped = any(order.has_company_debt_by_goods for order in orders)

        # Кредитный лимит
        credit_check = check_credit_limit(client)

        # Просроченные задолженности (продажи старше 30 дней с непогашенным долгом)
        from django.utils import timezone as tz
        import datetime
        threshold = tz.now().date() - datetime.timedelta(days=30)
        overdue_orders = [
            o for o in orders
            if o.has_company_debt_by_goods and o.date < threshold
        ]

        return Response({
            'client_id': client.id,
            'client_name': client.name,
            'orders': OrderSerializer(orders, many=True).data,
            'sales': SaleSerializer(sales, many=True).data,
            'payments': PaymentSerializer(payments, many=True).data,
            'returns': ReturnSerializer(returns, many=True).data,
            # Денежные итоги
            'total_revenue': api_decimal_str(total_revenue),
            'total_ordered': api_decimal_str(total_revenue),
            'total_paid': api_decimal_str(net_paid),
            'total_paid_gross': api_decimal_str(total_paid),
            'total_refunded': api_decimal_str(total_refunded),
            'client_debt_money': api_decimal_str(client_debt_money),
            'client_advance_amount': api_decimal_str(client_advance_amount),
            # Товарный долг
            'has_unshipped_goods': has_unshipped,
            'overdue_orders_count': len(overdue_orders),
            # Прибыль
            'total_profit': api_decimal_str(total_profit),
            'defect_revenue': api_decimal_str(defect_revenue),
            # Кредитный лимит
            'credit_limit': api_decimal_str(credit_check.credit_limit) if credit_check.credit_limit is not None else None,
            'credit_limit_mode': credit_check.block_mode,
            'credit_available': api_decimal_str(credit_check.credit_available) if credit_check.credit_available is not None else None,
            'credit_is_over_limit': credit_check.is_over_limit,
            'credit_warning': credit_check.warning,
        })

    @action(detail=True, methods=['get'], url_path='profile')
    def profile(self, request, pk=None):
        """
        Карточка клиента под продажи:
        - client info
        - total_debt
        - purchases (sales)
        - orders
        - returns
        """
        from .payment_status import sale_payment_metrics

        client = self.get_object()
        sales_qs = (
            Sale.objects.filter(client=client)
            .exclude(sale_status=Sale.STATUS_CANCELED)
            .order_by('-date', '-id')
        )
        orders_qs = Order.objects.filter(client=client).select_related('production_profile').order_by('-date', '-id')
        returns_qs = Return.objects.filter(sale__client=client).order_by('-date', '-id')

        total_debt = Decimal('0')
        total_sales_amount = Decimal('0')
        total_paid_amount = Decimal('0')
        debts = []
        sales_payload = []
        for s in sales_qs:
            m = sale_payment_metrics(s)
            sale_total = Decimal(str(s.revenue or 0)).quantize(Decimal('0.01'))
            paid = Decimal(str(m.get('paid_amount') or 0)).quantize(Decimal('0.01'))
            debt = Decimal(str(m.get('debt_amount') or 0)).quantize(Decimal('0.01'))
            total_sales_amount += sale_total
            total_paid_amount += paid
            total_debt += debt
            sale_item = {
                'id': s.id,
                'display': f'{s.sale_number or s.order_number} — {api_decimal_str(sale_total)}',
                'date': s.date.isoformat() if s.date else None,
                'sale_number': s.sale_number or s.order_number,
                'items': [
                    {
                        'product': sl.product,
                        'quantity': api_decimal_str(Decimal(str(sl.quantity or 0))),
                        'unit_price': api_decimal_str(Decimal(str(sl.unit_price or 0))),
                        'line_total': api_decimal_str(Decimal(str(sl.line_total or 0))),
                    }
                    for sl in s.sale_lines.all()
                ],
                'total_amount': api_decimal_str(sale_total),
                'paid_amount': api_decimal_str(paid),
                'debt_amount': api_decimal_str(debt),
                'payment_status': m.get('payment_status'),
                'payment_status_label': (
                    'Оплачено' if debt == 0 else ('Частично оплачено' if paid > 0 else 'В долг')
                ),
            }
            sales_payload.append(sale_item)
            if debt > 0:
                debts.append({
                    'id': s.id,
                    'sale_id': s.id,
                    'display': f'{s.sale_number or s.order_number} — долг {api_decimal_str(debt)}',
                    'sale_number': s.sale_number or s.order_number,
                    'date': s.date.isoformat() if s.date else None,
                    'total_amount': api_decimal_str(sale_total),
                    'paid_amount': api_decimal_str(paid),
                    'debt_amount': api_decimal_str(debt),
                })

        total_debt = sum((Decimal(str(d['debt_amount'])) for d in debts), Decimal('0'))
        orders_payload = [_order_display_payload(o) for o in orders_qs]
        returns_payload = []
        for ret in returns_qs:
            ret_label = ret.return_number or f'RET-{ret.id}'
            returns_payload.append({
                'id': ret.id,
                'display': f'Возврат {ret_label}',
                'date': ret.date.isoformat() if ret.date else None,
                'return_reason': ret.comment or '',
                'status': ret.status,
            })

        return Response({
            'client': {
                'id': client.id,
                'client_type': client.client_type or Client.TYPE_INDIVIDUAL,
                'name': client.name,
                'phone': client.phone or '',
                'phone_extra': client.phone_alt or '',
                'settlement_account': client.settlement_account or '',
                'inn': client.inn or '',
                'address': client.address or '',
                'status': 'active' if client.is_active else 'inactive',
                'status_label': 'Активен' if client.is_active else 'Неактивен',
            },
            'summary': {
                'total_sales_amount': api_decimal_str(total_sales_amount.quantize(Decimal('0.01'))),
                'total_paid_amount': api_decimal_str(total_paid_amount.quantize(Decimal('0.01'))),
                'total_debt': api_decimal_str(total_debt.quantize(Decimal('0.01'))),
                'total_orders': orders_qs.count(),
                'total_returns': returns_qs.count(),
            },
            'total_debt': api_decimal_str(total_debt.quantize(Decimal('0.01'))),
            'purchases': sales_payload,
            'sales': sales_payload,
            'orders': orders_payload,
            'returns': returns_payload,
            'debts': debts,
        })


# ─────────────────────────────────────────────────────────────────────────────
# ORDER (Заявка)
# ─────────────────────────────────────────────────────────────────────────────

_OrderCartLineWrite = inline_serializer(
    name='OrderCartLineWrite',
    fields={
        'profile': drf_serializers.IntegerField(help_text='ID пластикового профиля'),
        'quantity': drf_serializers.IntegerField(help_text='Количество (шт), > 0'),
        'length': drf_serializers.CharField(required=False, allow_null=True, allow_blank=True),
        'recipe': drf_serializers.IntegerField(required=False, allow_null=True),
    },
)
_OrderCreateRequestBody = inline_serializer(
    name='OrderCreateRequestBody',
    fields={
        'client': drf_serializers.IntegerField(),
        'date': drf_serializers.DateField(required=False),
        'order_lines': drf_serializers.ListField(child=_OrderCartLineWrite, min_length=1, required=False),
        'profile': drf_serializers.IntegerField(required=False),
        'quantity': drf_serializers.IntegerField(required=False),
        'length': drf_serializers.CharField(required=False, allow_null=True, allow_blank=True),
        'recipe': drf_serializers.IntegerField(required=False, allow_null=True),
        'total_amount': drf_serializers.CharField(required=False),
        'paid_amount': drf_serializers.CharField(required=False),
        'amount_remaining': drf_serializers.CharField(required=False),
        'payment_type': drf_serializers.ChoiceField(choices=['full', 'partial', 'debt']),
        'payment_method': drf_serializers.CharField(required=False, allow_blank=True, allow_null=True),
        'source_type': drf_serializers.CharField(required=False),
        'comment': drf_serializers.CharField(required=False, allow_blank=True),
    },
)


@extend_schema_view(
    create=extend_schema(
        summary='Создать заявку',
        description=(
            'Минимум по строкам: profile + quantity. Поля recipe и length не обязательны; '
            'либо массив order_lines, либо устаревший вариант с profile и quantity в корне. '
            'При нескольких активных рецептах у профиля без recipe в строке — 400 AMBIGUOUS_RECIPE_FOR_PROFILE.'
        ),
        request=_OrderCreateRequestBody,
        responses={201: OrderSerializer},
    ),
)
class OrderViewSet(ActivityLoggingMixin, viewsets.ModelViewSet):
    queryset = Order.objects.select_related(
        'client', 'created_by', 'responsible_user', 'resolved_recipe', 'production_profile',
    ).prefetch_related('lines', 'lines__profile', 'payments', 'sales').all()
    serializer_class = OrderSerializer
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'client_orders'
    activity_section = 'Заявки'
    activity_label = 'заявка'
    ordering_fields = ['id', 'date', 'status']
    filterset_class = OrderFilter

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(created_by=user)

    def perform_update(self, serializer):
        super().perform_update(serializer)

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        data = OrderSerializer(instance, context={'request': request}).data
        from .state_machine import ORDER_TRANSITIONS
        data['linked_entities'] = {
            'client': (
                {
                    'id': instance.client_id,
                    'label': instance.client.name,
                } if instance.client_id else None
            ),
            'responsible_user': (
                {
                    'id': instance.responsible_user_id,
                    'label': getattr(instance.responsible_user, 'name', '') or '',
                } if instance.responsible_user_id else None
            ),
        }
        data['available_status_transitions'] = ORDER_TRANSITIONS.get(instance.status, [])
        data['available_actions'] = {
            'set_status': bool(ORDER_TRANSITIONS.get(instance.status, [])),
            'reserve': instance.status in (
                Order.STATUS_NEW,
                Order.STATUS_CONFIRMED,
                Order.STATUS_IN_PROGRESS,
                Order.STATUS_PARTIALLY_SHIPPED,
            ),
            'release_reserve': True,
            'cancel': instance.status not in (Order.STATUS_CLOSED, Order.STATUS_CANCELED),
            'waybill': True,
            'history': True,
        }
        return Response(data)

    @action(detail=False, methods=['get'], url_path='select-sources')
    @extend_schema(
        description=(
            'Справочники для формы заявок: clients/profiles/recipes. '
            'recipes содержит profile_id для клиентской фильтрации без доп. запроса.'
        ),
    )
    def select_sources(self, request):
        from apps.recipes.models import PlasticProfile, Recipe

        clients = list(
            Client.objects.filter(is_active=True)
            .order_by('name')
            .values('id', 'name')[:200]
        )
        profiles = list(
            PlasticProfile.objects.order_by('name')
            .values('id', 'name')[:300]
        )
        recipes = list(
            Recipe.objects.select_related('profile')
            .order_by('recipe', 'id')
            .values('id', 'profile_id', 'recipe', 'product', 'is_active')[:1200]
        )
        return Response({
            'clients': [{'id': c['id'], 'label': c['name']} for c in clients],
            'profiles': [{'id': p['id'], 'label': p['name']} for p in profiles],
            'recipes': [
                {
                    'id': r['id'],
                    'profile_id': r['profile_id'],
                    'recipe': r['recipe'],
                    'name': (r['recipe'] or '').strip() or (r['product'] or ''),
                    'recipe_name': (r['recipe'] or '').strip() or (r['product'] or ''),
                    'is_active': bool(r['is_active']),
                }
                for r in recipes
            ],
        })

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        """
        Производство: draft → approved (коммит) → checking (коммит) → ready | not_ready (коммит),
        чтобы фронт и опросы видели «приняли» и «идёт проверка».
        """
        try:
            pre = Order.objects.get(pk=pk)
        except Order.DoesNotExist:
            return _err('NOT_FOUND', 'Заявка не найдена', http_status=404)
        if pre.request_status != Order.REQUEST_STATUS_DRAFT:
            return _err(
                'INVALID_REQUEST_STATUS',
                'Принять можно только заявку в статусе draft (производство).',
                http_status=400,
            )
        if not pre.production_profile_id or pre.production_length is None or not pre.production_quantity:
            return _err('INCOMPLETE_PRODUCTION', 'Укажите profile, length и quantity у заявки.', http_status=400)

        with transaction.atomic():
            order = Order.objects.select_for_update().get(pk=pk)
            if order.request_status != Order.REQUEST_STATUS_DRAFT:
                return _err(
                    'INVALID_REQUEST_STATUS',
                    'Принять можно только заявку в статусе draft (производство).',
                    http_status=400,
                )
            order.request_status = Order.REQUEST_STATUS_APPROVED
            order.save(update_fields=['request_status', 'updated_at'])

        with transaction.atomic():
            order = Order.objects.select_for_update().get(pk=pk)
            if order.request_status != Order.REQUEST_STATUS_APPROVED:
                return _err('INVALID_REQUEST_STATUS', 'Неконсистентный статус заявки (ожидался approved).', http_status=409)
            order.request_status = Order.REQUEST_STATUS_CHECKING
            order.save(update_fields=['request_status', 'updated_at'])

        with transaction.atomic():
            order = (
                Order.objects.select_for_update()
                .select_related('production_profile', 'client', 'resolved_recipe')
                .get(pk=pk)
            )
            if order.request_status != Order.REQUEST_STATUS_CHECKING:
                return _err('INVALID_REQUEST_STATUS', 'Неконсистентный статус заявки (ожидался checking).', http_status=409)
            apply_resource_check_to_order(order)
            order.save(
                update_fields=[
                    'resolved_recipe', 'resource_check_snapshot', 'request_status', 'updated_at',
                ],
            )
        return Response(OrderSerializer(order, context={'request': request}).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='recheck')
    def recheck(self, request, pk=None):
        """Повторная проверка ресурсов из not_ready: checking (коммит) → ready | not_ready."""
        try:
            pre = Order.objects.get(pk=pk)
        except Order.DoesNotExist:
            return _err('NOT_FOUND', 'Заявка не найдена', http_status=404)
        if pre.request_status != Order.REQUEST_STATUS_NOT_READY:
            return _err(
                'INVALID_REQUEST_STATUS',
                'Повторная проверка доступна только в статусе not_ready.',
                http_status=400,
            )
        if not pre.production_profile_id or pre.production_length is None or not pre.production_quantity:
            return _err('INCOMPLETE_PRODUCTION', 'Укажите profile, length и quantity у заявки.', http_status=400)

        with transaction.atomic():
            order = Order.objects.select_for_update().get(pk=pk)
            if order.request_status != Order.REQUEST_STATUS_NOT_READY:
                return _err(
                    'INVALID_REQUEST_STATUS',
                    'Повторная проверка доступна только в статусе not_ready.',
                    http_status=400,
                )
            order.request_status = Order.REQUEST_STATUS_CHECKING
            order.save(update_fields=['request_status', 'updated_at'])

        with transaction.atomic():
            order = (
                Order.objects.select_for_update()
                .select_related('production_profile', 'client', 'resolved_recipe')
                .get(pk=pk)
            )
            if order.request_status != Order.REQUEST_STATUS_CHECKING:
                return _err('INVALID_REQUEST_STATUS', 'Неконсистентный статус заявки (ожидался checking).', http_status=409)
            apply_resource_check_to_order(order)
            order.save(
                update_fields=[
                    'resolved_recipe', 'resource_check_snapshot', 'request_status', 'updated_at',
                ],
            )
        return Response(OrderSerializer(order, context={'request': request}).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        """Отклонить заявку (производство) → rejected."""
        with transaction.atomic():
            order = Order.objects.select_for_update().get(pk=pk)
            if order.request_status == Order.REQUEST_STATUS_IN_PRODUCTION:
                return _err(
                    'INVALID_REQUEST_STATUS',
                    'Нельзя отклонить заявку, уже запущенную в производство.',
                    http_status=400,
                )
            if order.request_status is None:
                return _err('INVALID_REQUEST_STATUS', 'Заявка не в цепочке производства (request_status).', http_status=400)
            if order.request_status == Order.REQUEST_STATUS_REJECTED:
                return Response(OrderSerializer(order, context={'request': request}).data, status=status.HTTP_200_OK)
            order.request_status = Order.REQUEST_STATUS_REJECTED
            order.save(update_fields=['request_status', 'updated_at'])
        return Response(OrderSerializer(order, context={'request': request}).data, status=status.HTTP_200_OK)

    @action(detail=True, methods=['patch'], url_path='status')
    def set_status(self, request, pk=None):
        """Изменить статус заявки с валидацией переходов (централизованный state machine)."""
        from .state_machine import validate_order_transition

        order = self.get_object()
        new_status = request.data.get('status')
        if not new_status:
            return _err('MISSING_STATUS', 'Укажите status')

        try:
            validate_order_transition(order.status, new_status)
        except ValueError as e:
            return _err('INVALID_STATUS_TRANSITION', str(e), http_status=422)

        if new_status in (
            Order.STATUS_SHIPPED,
            Order.STATUS_PARTIALLY_SHIPPED,
            Order.STATUS_CLOSED,
        ):
            from .order_sync import validate_order_for_new_status
            try:
                validate_order_for_new_status(order, new_status)
            except ValueError as e:
                return _err('ORDER_STATUS_BLOCKED', str(e), http_status=422)

        order.status = new_status
        order.save(update_fields=['status', 'updated_at'])
        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=['get'], url_path='waybill')
    def waybill(self, request, pk=None):
        """HTML-накладная по заявке (канонический endpoint)."""
        order = self.get_object()
        return _order_html_response(order)

    def destroy(self, request, *args, **kwargs):
        return Response(
            {
                'code': 'DELETE_DISABLED',
                'error': 'Физическое удаление заявок отключено. Используйте /api/orders/{id}/cancel/.',
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    @action(detail=True, methods=['get'], url_path='history')
    def order_history(self, request, pk=None):
        """Трассировка: заявка → продажи → возвраты → переделки."""
        order = self.get_object()
        sales = Sale.objects.filter(linked_order=order).prefetch_related('sale_lines')
        payments = Payment.objects.filter(linked_order=order, status=Payment.STATUS_ACTIVE)
        returns = Return.objects.filter(linked_order=order).prefetch_related('lines')
        return Response({
            'order': OrderSerializer(order).data,
            'sales': SaleSerializer(sales, many=True).data,
            'payments': PaymentSerializer(payments, many=True).data,
            'returns': ReturnSerializer(returns, many=True).data,
        })

    @action(detail=True, methods=['post'], url_path='reserve')
    def reserve(self, request, pk=None):
        """
        Зарезервировать товар под строку заявки.
        Body: { order_line_id, warehouse_batch_id, quantity, comment? }
        """
        from .reservations import reserve_order_line
        from apps.warehouse.models import WarehouseBatch

        order = self.get_object()
        line_id = request.data.get('order_line_id')
        wb_id = request.data.get('warehouse_batch_id')
        quantity_raw = request.data.get('quantity')
        comment = request.data.get('comment', '')

        if not line_id:
            return _err('MISSING_FIELD', 'Укажите order_line_id')
        if not wb_id:
            return _err('MISSING_FIELD', 'Укажите warehouse_batch_id')
        if quantity_raw is None:
            return _err('MISSING_FIELD', 'Укажите quantity')

        try:
            line = order.lines.get(pk=line_id)
        except OrderLine.DoesNotExist:
            return _err('NOT_FOUND', 'Строка заявки не найдена', http_status=404)

        try:
            wb = WarehouseBatch.objects.get(pk=wb_id)
        except WarehouseBatch.DoesNotExist:
            return _err('NOT_FOUND', 'Партия склада не найдена', http_status=404)

        try:
            qty = Decimal(str(quantity_raw))
        except Exception:
            return _err('INVALID_FIELD', 'Некорректное значение quantity')

        try:
            reservation = reserve_order_line(
                order_line=line,
                warehouse_batch=wb,
                quantity=qty,
                user=request.user if request.user.is_authenticated else None,
                comment=comment,
            )
        except ValueError as e:
            return _err('RESERVATION_ERROR', str(e), http_status=422)

        return Response(OrderReservationSerializer(reservation).data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='release-reserve')
    def release_reserve(self, request, pk=None):
        """
        Снять резерв.
        Body: { reservation_id }
        """
        from .reservations import release_reservation

        order = self.get_object()
        res_id = request.data.get('reservation_id')
        if not res_id:
            return _err('MISSING_FIELD', 'Укажите reservation_id')

        try:
            reservation = OrderReservation.objects.get(
                pk=res_id,
                order_line__order=order,
            )
        except OrderReservation.DoesNotExist:
            return _err('NOT_FOUND', 'Резерв не найден', http_status=404)

        try:
            reservation = release_reservation(reservation)
        except ValueError as e:
            return _err('RESERVATION_ERROR', str(e), http_status=422)

        return Response(OrderReservationSerializer(reservation).data)

    @action(detail=True, methods=['get'], url_path='reservations')
    def reservations(self, request, pk=None):
        """Список всех резервов по заявке."""
        order = self.get_object()
        line_ids = list(order.lines.values_list('id', flat=True))
        qs = OrderReservation.objects.filter(
            order_line_id__in=line_ids,
        ).select_related('order_line', 'warehouse_batch', 'created_by').order_by('-created_at')
        return Response(OrderReservationSerializer(qs, many=True).data)

    @action(detail=True, methods=['patch'], url_path='cancel')
    def cancel_order(self, request, pk=None):
        """
        Отменить заявку с автоматическим снятием всех активных резервов.
        """
        from .reservations import release_all_for_order
        from .state_machine import validate_order_transition, validate_order_cancel

        order = self.get_object()
        try:
            validate_order_transition(order.status, Order.STATUS_CANCELED)
            validate_order_cancel(order)
        except ValueError as e:
            return _err('INVALID_TRANSITION', str(e), http_status=422)

        released = release_all_for_order(order)
        order.status = Order.STATUS_CANCELED
        order.save(update_fields=['status', 'updated_at'])
        return Response({
            'status': order.status,
            'reservations_released': released,
            'order': OrderSerializer(order).data,
        })


def _order_html_response(order: Order) -> HttpResponse:
    parts = [
        '<!DOCTYPE html><html lang="ru"><head><meta charset="utf-8">',
        f'<title>Заявка {escape(order.order_number)}</title>',
        '<style>body{font-family:system-ui,sans-serif;max-width:800px;margin:2rem auto;line-height:1.5;}',
        'table{border-collapse:collapse;width:100%;}td,th{border:1px solid #ccc;padding:0.4rem 0.6rem;text-align:left;}',
        'th{background:#f5f5f5;}</style>',
        '</head><body>',
        '<h1>Заявка</h1>',
        f'<p><strong>№</strong> {escape(order.order_number)} <strong>от</strong> {order.date.isoformat()}</p>',
        f'<p><strong>Статус:</strong> {escape(order.get_status_display())}</p>',
    ]
    if order.client_id:
        cl = order.client
        parts.append(f'<p><strong>Покупатель:</strong> {escape(cl.name)}</p>')
        if cl.inn:
            parts.append(f'<p>ИНН: {escape(cl.inn)}</p>')
        if cl.phone:
            parts.append(f'<p>Тел.: {escape(cl.phone)}</p>')
    parts.append(
        '<table><thead><tr>'
        '<th>№</th><th>Наименование</th><th>Тип</th>'
        '<th>Заказано</th><th>Отгружено</th><th>Остаток</th>'
        '<th>Цена</th><th>Сумма</th>'
        '</tr></thead><tbody>'
    )
    for i, line in enumerate(order.lines.all(), 1):
        price_str = str(line.unit_price) if line.unit_price is not None else '—'
        total_str = str(line.line_total) if line.unit_price is not None else '—'
        parts.append(
            f'<tr>'
            f'<td>{i}</td>'
            f'<td>{escape(line.product)}</td>'
            f'<td>{escape(line.product_type or "")}</td>'
            f'<td>{escape(str(line.ordered_quantity))}</td>'
            f'<td>{escape(str(line.shipped_quantity))}</td>'
            f'<td>{escape(str(line.remaining_quantity))}</td>'
            f'<td>{escape(price_str)}</td>'
            f'<td>{escape(total_str)}</td>'
            f'</tr>'
        )
    parts.append('</tbody></table>')
    if order.comment:
        parts.append(f'<p><strong>Комментарий:</strong> {escape(order.comment)}</p>')
    parts.append('<p><em>Сформировано автоматически.</em></p></body></html>')
    html = ''.join(parts)
    resp = HttpResponse(html, content_type='text/html; charset=utf-8')
    resp['Content-Disposition'] = f'inline; filename="order-waybill-{order.id}.html"'
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# SALE (Продажа)
# ─────────────────────────────────────────────────────────────────────────────

class SaleViewSet(ActivityLoggingMixin, viewsets.ModelViewSet):
    queryset = Sale.objects.select_related(
        'client', 'warehouse_batch', 'warehouse_batch__profile', 'linked_order',
    ).prefetch_related('sale_lines', 'payments').all()
    serializer_class = SaleSerializer
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'sales'
    activity_section = 'Продажи'
    activity_label = 'продажа'
    filterset_class = SaleFilter
    ordering_fields = ['id', 'date']
    format_kwarg = None

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(created_by=user)

    def perform_update(self, serializer):
        instance = serializer.instance
        if instance.warehouse_batch_id:
            if 'warehouse_batch' in serializer.validated_data:
                wb = serializer.validated_data['warehouse_batch']
                new_id = wb.pk if wb else None
                if new_id is None:
                    raise ValidationError(
                        {'warehouse_batch': 'Нельзя отвязать партию склада у продажи'},
                    )
                if new_id != instance.warehouse_batch_id:
                    raise ValidationError(
                        {'warehouse_batch': 'Нельзя сменить партию склада у существующей продажи'},
                    )
            if 'quantity' in serializer.validated_data:
                new_q = serializer.validated_data['quantity']
                if Decimal(str(new_q)) != Decimal(str(instance.quantity)):
                    raise ValidationError(
                        {'quantity': 'Нельзя изменить количество: создайте новую продажу или отмените эту'},
                    )
            if 'quantity_input' in serializer.validated_data:
                new_qi = serializer.validated_data.get('quantity_input')
                old_qi = instance.quantity_input
                if new_qi is None and old_qi is None:
                    pass
                elif new_qi is None or old_qi is None:
                    raise ValidationError(
                        {'quantity_input': 'Нельзя изменить quantity_input: создайте новую продажу или отмените эту'},
                    )
                elif Decimal(str(new_qi)) != Decimal(str(old_qi)):
                    raise ValidationError(
                        {'quantity_input': 'Нельзя изменить quantity_input: создайте новую продажу или отмените эту'},
                    )
            if 'stock_form' in serializer.validated_data or 'piece_pick' in serializer.validated_data:
                raise ValidationError(
                    {'stock_form': 'Нельзя менять stock_form / piece_pick после создания продажи со складом'},
                )
        super().perform_update(serializer)

    def destroy(self, request, *args, **kwargs):
        return Response(
            {'code': 'DELETE_DISABLED', 'error': 'Удаление продажи отключено. Используйте /api/sales/{id}/cancel/.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        data = SaleSerializer(instance, context={'request': request}).data
        from .state_machine import SALE_TRANSITIONS
        from .payment_status import sale_payment_metrics
        pay = sale_payment_metrics(instance)
        paid_amount = Decimal(str(pay.get('paid_amount') or 0)).quantize(Decimal('0.01'))
        debt_amount = Decimal(str(pay.get('debt_amount') or 0)).quantize(Decimal('0.01'))
        payment_status = pay.get('payment_status') or ('paid' if debt_amount == 0 else ('partial' if paid_amount > 0 else 'debt'))

        active_payments = list(
            instance.payments.filter(
                status=Payment.STATUS_ACTIVE,
                payment_type__in=(Payment.TYPE_PAYMENT, Payment.TYPE_PREPAYMENT, Payment.TYPE_SURCHARGE),
            ).order_by('-id')
        )
        payment_method_raw = active_payments[0].payment_method if active_payments else None
        payment_method = (
            'cash' if payment_method_raw == Payment.METHOD_CASH
            else ('card' if payment_method_raw == Payment.METHOD_CARD
            else ('transfer' if payment_method_raw == Payment.METHOD_TRANSFER else None))
        )
        if debt_amount == 0:
            payment_type = 'full'
        elif paid_amount > 0:
            payment_type = 'partial'
        else:
            payment_type = 'debt'
        payment_type_label = {'full': 'Полная оплата', 'partial': 'Частичная оплата', 'debt': 'В долг'}[payment_type]
        payment_method_label = {'cash': 'Наличные', 'card': 'Карта', 'transfer': 'Перевод'}.get(payment_method, '')
        payment_status_label = {
            'paid': 'Оплачено',
            'unpaid': 'Долг',
            'debt': 'Долг',
            'partially_paid': 'Частично оплачено',
            'partial': 'Частично оплачено',
            'overpaid': 'Переплата',
            'refunded': 'Возвращено',
        }.get(payment_status, '—')

        sale_lines = []
        all_sale_lines = list(instance.sale_lines.select_related('warehouse_batch', 'warehouse_batch__profile').all())
        lines_count = len(all_sale_lines)
        for sl in all_sale_lines:
            wb = sl.warehouse_batch
            ol = sl.order_line
            wb_display = None
            if wb is not None:
                wb_display = (
                    f"{(wb.profile.name if wb.profile_id else wb.product)} — "
                    f"{api_decimal_str(Decimal(str(wb.length_per_piece or 0)))} м"
                )
            line_display = wb_display or sl.product
            pkg = _sale_line_packaging_payload(instance, sl, lines_count=lines_count)
            sale_lines.append({
                'id': sl.id,
                'order_line': sl.order_line_id,
                'order_line_id': sl.order_line_id,
                'order_line_display': (
                    f"{ol.product} — заказано {api_decimal_str(Decimal(str(ol.ordered_quantity or 0)))} — "
                    f"отгружено {api_decimal_str(Decimal(str(ol.shipped_quantity or 0)))} — "
                    f"осталось {api_decimal_str(ol.remaining_quantity)}"
                ) if ol is not None else None,
                'order_line_product': ol.product if ol is not None else None,
                'unit_type': pkg['unit_type'],
                'warehouse_batch_display': wb_display,
                'warehouse_batch': sl.warehouse_batch_id,
                'display': line_display,
                'quantity': pkg['quantity'],
                'packages_quantity': pkg['packages_quantity'],
                'pieces_per_package': pkg['pieces_per_package'],
                'quantity_display': pkg['quantity_display'],
                'unit_price': api_decimal_str(Decimal(str(sl.unit_price or 0))),
                'total_amount': api_decimal_str(Decimal(str(sl.line_total or 0))),
                'line_total': api_decimal_str(Decimal(str(sl.line_total or 0))),
            })

        data['order_display'] = (
            _order_display_payload(instance.linked_order)['display']
            if instance.linked_order_id else None
        )
        data['payment_type'] = payment_type
        data['payment_type_label'] = payment_type_label
        data['payment_method'] = payment_method
        data['payment_method_label'] = payment_method_label
        data['payment_status'] = payment_status
        data['payment_status_label'] = payment_status_label
        data['total_amount'] = api_decimal_str(Decimal(str(instance.revenue or 0)))
        data['paid_amount'] = api_decimal_str(paid_amount)
        data['debt_amount'] = api_decimal_str(debt_amount)
        data['sale_lines'] = sale_lines
        data['unit_type'] = instance.sale_mode
        data['linked_entities'] = {
            'client': (
                {
                    'id': instance.client_id,
                    'label': instance.client.name,
                } if instance.client_id else None
            ),
            'linked_order': (
                {
                    'id': instance.linked_order_id,
                    'label': instance.linked_order.order_number,
                } if instance.linked_order_id else None
            ),
            'warehouse_batch': (
                {
                    'id': instance.warehouse_batch_id,
                    'label': f'#{instance.warehouse_batch_id} {instance.warehouse_batch.product}',
                } if instance.warehouse_batch_id else None
            ),
        }
        data['available_status_transitions'] = SALE_TRANSITIONS.get(instance.sale_status, [])
        data['available_actions'] = {
            'set_status': bool(SALE_TRANSITIONS.get(instance.sale_status, [])),
            'credit_check': bool(instance.client_id),
            'waybill': True,
            'receipt': True,
        }
        return Response(data)

    @action(detail=False, methods=['get'], url_path='select-sources')
    @extend_schema(
        parameters=[
            OpenApiParameter(name='client', required=False, type=int),
            OpenApiParameter(name='client_id', required=False, type=int),
            OpenApiParameter(name='order', required=False, type=int),
            OpenApiParameter(name='order_id', required=False, type=int),
            OpenApiParameter(name='unit_type', required=False, type=str),
        ],
        description=(
            'Справочники для формы продаж. В каждом order в полях orders/available_orders '
            'возвращаются paid_amount/debt_amount/total_amount/payment_type/payment_method.\n'
            'Партии warehouse_batches / available_warehouse_batches:\n'
            '- available_pieces_total — суммарные доступные штуки по партии (get_available_quantity).\n'
            '- available_pieces, available_unpacked_pieces, unpacked_pieces — только неупакованный хвост; '
            'для packed всегда 0 (штуки только в упаковках / GP).\n'
            '- available_packages — только когда известны запечатанные упаковки: для packed = total/ipp; '
            'для unpacked/open_package — только если задано packages_count>0 (иначе null, '
            'чтобы фронт не вычитал «виртуальные коробки» из total при отсутствии учёта запечатанного).\n'
            'При unit_type=pieces партии packed не включаются; строки без неупакованного хвоста '
            '(unpacked_pieces<=0) тоже исключаются. '
            'supports_packages=true при inventory_form=packed и pieces_per_package>0. '
            'Строки GP после POST /warehouse/gp-packages/ попадают сюда как отдельные packed-партии (1 упаковка = 1 warehouse_batch).'
        ),
    )
    def select_sources(self, request):
        from apps.warehouse.models import WarehouseBatch
        from .reservations import get_available_quantity

        clients = list(
            Client.objects.filter(is_active=True)
            .order_by('name')
            .values('id', 'name')[:200]
        )
        orders_qs = Order.objects.order_by('-date', '-id')
        client_id = request.query_params.get('client_id')
        if not client_id:
            client_id = request.query_params.get('client')
        unit_type = (request.query_params.get('unit_type') or request.query_params.get('sale_mode') or '').strip().lower()
        if client_id:
            orders_qs = orders_qs.filter(client_id=client_id)
        orders = list(orders_qs.select_related('client', 'production_profile', 'resolved_recipe')[:200])
        order_id = request.query_params.get('order_id') or request.query_params.get('order')
        order_lines = []
        if order_id:
            lines_qs = OrderLine.objects.filter(order_id=order_id).order_by('id')[:300]
            order_lines = [
                {
                    'id': line.id,
                    'label': (
                        f"{line.product} — заказано {api_decimal_str(line.ordered_quantity)} — "
                        f"продано {api_decimal_str(line.shipped_quantity)} — "
                        f"осталось {api_decimal_str(line.remaining_quantity)}"
                    ),
                    'product': line.product,
                    'ordered_quantity': api_decimal_str(line.ordered_quantity),
                    'shipped_quantity': api_decimal_str(line.shipped_quantity),
                    'remaining_quantity': api_decimal_str(line.remaining_quantity),
                    'unit_price': api_decimal_str(line.unit_price or Decimal('0')),
                }
                for line in lines_qs
            ]
        wb_qs = (
            WarehouseBatch.objects.filter(
                status=WarehouseBatch.STATUS_AVAILABLE,
                quality=WarehouseBatch.QUALITY_GOOD,
                stock_bucket=WarehouseBatch.STOCK_BUCKET_STANDARD,
            )
            .order_by('-date', '-id')
            [:300]
        )
        warehouse_batches = []
        available_batches = []
        quality_labels = {
            WarehouseBatch.QUALITY_GOOD: 'Годный',
            WarehouseBatch.QUALITY_DEFECT: 'Брак',
        }
        inventory_labels = {
            WarehouseBatch.INVENTORY_PACKED: 'Упаковано',
            WarehouseBatch.INVENTORY_OPEN_PACKAGE: 'Открытая упаковка',
            WarehouseBatch.INVENTORY_UNPACKED: 'Неупаковано',
        }
        for b in wb_qs:
            raw_available = Decimal(str(get_available_quantity(b.pk)))
            if raw_available <= 0:
                continue
            ppp = Decimal(str(b.pieces_per_package or 0)) if b.pieces_per_package is not None else Decimal('0')
            av = _warehouse_batch_select_sources_availability(b, raw_available)
            loose_available = av['unpacked_pieces']
            total_pieces = av['total_pieces']
            avail_packages = av['available_packages_str']
            is_packed = b.inventory_form == WarehouseBatch.INVENTORY_PACKED
            # Во вкладке «штуки» не показываем закрытые упаковки — остаток только в GP/упаковках.
            if unit_type == Sale.MODE_PACKAGES:
                if not is_packed:
                    continue
                if ppp <= 0:
                    continue
                if (raw_available / ppp) < Decimal('1'):
                    continue
            elif unit_type == Sale.MODE_PIECES:
                if is_packed:
                    continue
                if loose_available <= 0:
                    continue

            total_meters = None
            if b.length_per_piece is not None:
                total_meters = api_decimal_str(
                    (raw_available * Decimal(str(b.length_per_piece))).quantize(Decimal('0.0001')),
                )
            if is_packed:
                display = (
                    f"{(b.profile.name if b.profile_id else b.product)} — "
                    f"{api_decimal_str(Decimal(str(b.length_per_piece or 0)))} м — "
                    f"остаток: {avail_packages or '0'} уп."
                )
            else:
                display = (
                    f"{(b.profile.name if b.profile_id else b.product)} — "
                    f"{api_decimal_str(Decimal(str(b.length_per_piece or 0)))} м — "
                    f"остаток: {api_decimal_str(loose_available)} шт"
                )
                if avail_packages is not None:
                    display += f" / {avail_packages} уп."
            if is_packed and ppp > 0 and avail_packages is not None:
                free_label = f'{avail_packages} уп.'
            else:
                free_label = f'{api_decimal_str(loose_available)} шт'
            loose_s = api_decimal_str(loose_available)
            total_s = api_decimal_str(total_pieces)
            warehouse_batches.append(
                {
                    'id': b.pk,
                    'label': (
                        f"#{b.pk} — {b.product} — свободно {free_label} — "
                        f"{quality_labels.get(b.quality, b.quality)} — "
                        f"{inventory_labels.get(b.inventory_form, b.inventory_form)}"
                    ),
                    'product': b.product,
                    'available_quantity': loose_s,
                    'available_pieces': loose_s,
                    'available_pieces_total': total_s,
                    'available_unpacked_pieces': loose_s,
                    'unpacked_pieces': loose_s,
                    'available_packages': avail_packages,
                    'supports_pieces': not is_packed,
                    'supports_packages': bool(
                        b.inventory_form == WarehouseBatch.INVENTORY_PACKED and ppp > 0
                    ),
                    'quality': b.quality,
                    'status': b.status,
                    'inventory_form': b.inventory_form,
                },
            )
            available_batches.append(
                {
                    'id': b.pk,
                    'display': display,
                    'warehouse_batch_display': display,
                    'profile_id': b.profile_id,
                    'profile_name': b.profile.name if b.profile_id else None,
                    'pieces_per_package': (
                        api_decimal_str(Decimal(str(b.pieces_per_package)))
                        if b.pieces_per_package is not None else None
                    ),
                    'length_per_piece': (
                        api_decimal_str(Decimal(str(b.length_per_piece)))
                        if b.length_per_piece is not None else None
                    ),
                    'available_pieces': loose_s,
                    'available_pieces_total': total_s,
                    'available_unpacked_pieces': loose_s,
                    'unpacked_pieces': loose_s,
                    'available_packages': avail_packages,
                    'supports_pieces': not is_packed,
                    'supports_packages': bool(
                        b.inventory_form == WarehouseBatch.INVENTORY_PACKED and ppp > 0
                    ),
                    'total_meters': total_meters,
                    'quality': b.quality,
                    'status': b.status,
                    'unit_labels': {'pieces': 'шт', 'packages': 'уп', 'meters': 'м'},
                },
            )
        available_orders = [_order_display_payload(o) for o in orders]
        orders_payload = []
        for o, op in zip(orders, available_orders):
            orders_payload.append(
                {
                    'id': o.id,
                    'client_id': o.client_id,
                    'label': (
                        f"{o.order_number} — {(o.client.name if o.client_id else '—')} — "
                        f"осталось {api_decimal_str(o.remaining_amount)}"
                    ),
                    'display': op.get('display'),
                    'order_display': op.get('order_display'),
                    'paid_amount': op['paid_amount'],
                    'prepaid_amount': op.get('prepayment_amount'),
                    'debt_amount': op['debt_amount'],
                    'total_amount': op['total_amount'],
                    'payment_type': op['payment_type'],
                    'payment_method': op['payment_method'],
                    'prepayment_amount': op['prepayment_amount'],
                    'advance_amount': op['advance_amount'],
                },
            )
        return Response({
            'clients': [{'id': c['id'], 'label': c['name']} for c in clients],
            'orders': orders_payload,
            'order_lines': order_lines,
            'warehouse_batches': warehouse_batches,
            'available_orders': available_orders,
            'available_warehouse_batches': available_batches,
        })

    @action(detail=False, methods=['post'], url_path='preview')
    def preview(self, request):
        """
        Предпросмотр продажи без списаний и без сохранения.
        """
        from apps.warehouse.models import WarehouseBatch
        from .reservations import get_available_quantity

        data = request.data or {}
        client_id = data.get('client')
        if client_id in (None, ''):
            client_id = data.get('client_id')
        if client_id in (None, ''):
            return _err('MISSING_CLIENT', 'Поле client (или client_id) обязательно.', http_status=400)
        sale_lines = data.get('sale_lines')
        if not isinstance(sale_lines, list) or len(sale_lines) < 1:
            return _err('MISSING_SALE_LINES', 'sale_lines обязателен и должен содержать минимум одну строку.', http_status=400)

        unit_type = (data.get('unit_type') or Sale.MODE_PIECES).strip().lower()
        if unit_type not in (Sale.MODE_PIECES, Sale.MODE_PACKAGES):
            return _err('INVALID_UNIT_TYPE', 'unit_type: pieces или packages', http_status=400)

        total_amount = Decimal('0')
        normalized_lines = []
        errors = []
        for idx, row in enumerate(sale_lines, start=1):
            line_unit_type = (row.get('unit_type') or unit_type or Sale.MODE_PIECES)
            line_unit_type = str(line_unit_type).strip().lower()
            if line_unit_type not in (Sale.MODE_PIECES, Sale.MODE_PACKAGES):
                errors.append(
                    {
                        'field': f'sale_lines[{idx}].unit_type',
                        'message': 'sale_lines[].unit_type: pieces | packages',
                    },
                )
                continue
            wb_id = row.get('warehouse_batch')
            qty_raw = row.get('quantity')
            up_raw = row.get('unit_price')
            if wb_id in (None, ''):
                errors.append({'field': f'sale_lines[{idx}].warehouse_batch', 'message': 'warehouse_batch обязателен'})
                continue
            if qty_raw in (None, ''):
                errors.append({'field': f'sale_lines[{idx}].quantity', 'message': 'quantity обязателен'})
                continue
            if up_raw in (None, ''):
                errors.append({'field': f'sale_lines[{idx}].unit_price', 'message': 'unit_price обязателен'})
                continue
            try:
                wb = WarehouseBatch.objects.select_related('profile').get(pk=wb_id)
            except WarehouseBatch.DoesNotExist:
                errors.append({'field': f'sale_lines[{idx}].warehouse_batch', 'message': 'Партия не найдена'})
                continue
            try:
                qty_in = Decimal(str(qty_raw))
            except (InvalidOperation, TypeError, ValueError):
                errors.append({'field': f'sale_lines[{idx}].quantity', 'message': 'quantity должен быть числом'})
                continue
            try:
                unit_price = Decimal(str(up_raw))
            except (InvalidOperation, TypeError, ValueError):
                errors.append({'field': f'sale_lines[{idx}].unit_price', 'message': 'unit_price должен быть числом'})
                continue
            if qty_in <= 0:
                errors.append({'field': f'sale_lines[{idx}].quantity', 'message': 'quantity должен быть > 0'})
                continue
            if unit_price < 0:
                errors.append({'field': f'sale_lines[{idx}].unit_price', 'message': 'unit_price не может быть < 0'})
                continue
            available_pieces = Decimal(str(get_available_quantity(wb.pk)))
            if line_unit_type == Sale.MODE_PACKAGES:
                try:
                    ppp = Decimal(str(wb.pieces_per_package or 0))
                except (InvalidOperation, TypeError, ValueError):
                    ppp = Decimal('0')
                if wb.inventory_form != WarehouseBatch.INVENTORY_PACKED:
                    errors.append(
                        {
                            'field': f'sale_lines[{idx}].warehouse_batch',
                            'message': 'Для unit_type=packages нужна партия с inventory_form=packed',
                        },
                    )
                    continue
                if ppp <= 0:
                    errors.append({'field': f'sale_lines[{idx}].quantity', 'message': 'Для продажи в упаковках у партии нет pieces_per_package'})
                    continue
                qty_pieces = (qty_in * ppp).quantize(Decimal('0.0001'))
            else:
                qty_pieces = qty_in.quantize(Decimal('0.0001'))
            if qty_pieces > available_pieces + Decimal('0.0001'):
                errors.append({
                    'field': f'sale_lines[{idx}].quantity',
                    'message': f'Недостаточно остатка: доступно {api_decimal_str(available_pieces)} шт',
                })
                continue
            line_total = (qty_in * unit_price).quantize(Decimal('0.01'))
            total_amount += line_total
            normalized_lines.append({
                'warehouse_batch': wb.pk,
                'warehouse_batch_display': f"{(wb.profile.name if wb.profile_id else wb.product)}",
                'unit_type': line_unit_type,
                'input_quantity': api_decimal_str(qty_in),
                'quantity_pieces': api_decimal_str(qty_pieces),
                'unit_price': api_decimal_str(unit_price),
                'line_total': api_decimal_str(line_total),
            })

        if errors:
            return _err('VALIDATION_ERROR', 'Ошибки в предпросмотре продажи.', errors=errors, http_status=400)

        payment_type = (data.get('payment_type') or '').strip().lower()
        payment_method = (data.get('payment_method') or '').strip().lower()
        paid_raw = data.get('paid_amount')
        if not payment_type:
            if paid_raw not in (None, ''):
                try:
                    pd = Decimal(str(paid_raw))
                    payment_type = 'partial' if pd > 0 else 'debt'
                except (InvalidOperation, TypeError, ValueError):
                    payment_type = 'debt'
                    paid_raw = None
            else:
                payment_type = 'debt'
        if not payment_method:
            payment_method = 'cash'
        total_amount = total_amount.quantize(Decimal('0.01'))
        paid_amount = Decimal('0')
        debt_amount = Decimal('0')
        if payment_type == 'full':
            paid_amount = total_amount
            debt_amount = Decimal('0')
        elif payment_type == 'partial':
            if paid_raw in (None, ''):
                return _err('PAID_AMOUNT_REQUIRED', 'Для partial укажите paid_amount.', http_status=400)
            try:
                paid_amount = Decimal(str(paid_raw)).quantize(Decimal('0.01'))
            except (InvalidOperation, TypeError, ValueError):
                return _err('INVALID_PAID_AMOUNT', 'paid_amount должен быть числом.', http_status=400)
            if paid_amount <= 0 or paid_amount > total_amount:
                return _err('INVALID_PAID_AMOUNT', 'paid_amount должен быть в диапазоне (0, total_amount].', http_status=400)
            debt_amount = (total_amount - paid_amount).quantize(Decimal('0.01'))
        elif payment_type == 'debt':
            paid_amount = Decimal('0')
            debt_amount = total_amount
        else:
            return _err('INVALID_PAYMENT_TYPE', 'payment_type: full | partial | debt', http_status=400)

        if payment_method not in ('cash', 'card', 'transfer'):
            return _err('INVALID_PAYMENT_METHOD', 'payment_method: cash | card | transfer', http_status=400)

        payment_status = 'paid' if debt_amount == 0 else ('partial' if paid_amount > 0 else 'debt')
        return Response({
            'total_amount': api_decimal_str(total_amount),
            'paid_amount': api_decimal_str(paid_amount),
            'debt_amount': api_decimal_str(debt_amount),
            'payment_status': payment_status,
            'payment_type_label': {'full': 'Полная оплата', 'partial': 'Частичная оплата', 'debt': 'В долг'}[payment_type],
            'payment_method_label': {'cash': 'Наличные', 'card': 'Карта', 'transfer': 'Перевод'}[payment_method],
            'payment_status_label': {'paid': 'Оплачено', 'partial': 'Частично оплачено', 'debt': 'В долг'}[payment_status],
            'summary': (
                f"Итого {api_decimal_str(total_amount)}; оплачено {api_decimal_str(paid_amount)}; "
                f"долг {api_decimal_str(debt_amount)}"
            ),
            'unit_type': unit_type,
            'normalized_lines': normalized_lines,
        })

    @staticmethod
    def _nakladnaya_html_response(sale):
        payload = _build_sale_waybill_payload(sale)
        html = _render_waybill_html(payload)
        resp = HttpResponse(html, content_type='text/html; charset=utf-8')
        resp['Content-Disposition'] = f'inline; filename="waybill-{sale.id}.html"'
        return resp

    def _serve_nakladnaya(self, request, *args, **kwargs):
        accept = (request.headers.get('Accept') or '').lower()
        requested_format = (request.query_params.get('format') or '').strip().lower()
        if requested_format:
            fmt = requested_format
        elif 'application/json' in accept:
            fmt = 'json'
        else:
            fmt = 'html'
        if fmt not in ('html', 'pdf', 'xlsx', 'json'):
            return _err('INVALID_FORMAT', 'Поддерживаемые форматы: html, pdf, xlsx, json', http_status=400)
        try:
            sale = self.get_object()
        except Http404:
            return _err('SALE_NOT_FOUND', 'Продажа не найдена', http_status=404)

        payload = _build_sale_waybill_payload(sale)
        payload['title'] = f"Расходная накладная № {sale.id} от ________ г."
        payload['buyer_name'] = sale.client.name if sale.client_id else '—'

        if fmt == 'json':
            return Response(_waybill_json_payload(payload))
        if fmt == 'html':
            return SaleViewSet._nakladnaya_html_response(sale)
        if fmt == 'pdf':
            return _sale_waybill_pdf_response(payload)
        return _sale_waybill_xlsx_response(payload)

    @action(detail=True, methods=['post', 'patch'], url_path='cancel')
    def cancel_sale(self, request, pk=None):
        """Отмена продажи: откат склада, восстановление резервов, статус canceled."""
        from .state_machine import validate_sale_transition
        from .sale_warehouse import reverse_warehouse_for_sale
        from .reservations import restore_reservations_for_sale

        sale = self.get_object()
        if sale.sale_status == Sale.STATUS_CANCELED:
            return _err('ALREADY_CANCELED', 'Продажа уже отменена', http_status=422)
        if Return.objects.filter(sale=sale).exclude(status=Return.STATUS_CANCELED).exists():
            return _err('HAS_RETURNS', 'Нельзя отменить продажу: есть возвраты. Обработайте возвраты вручную.', http_status=409)
        if Payment.objects.filter(linked_sale=sale, status=Payment.STATUS_ACTIVE).exists():
            return _err('HAS_PAYMENTS', 'Нельзя отменить продажу: есть оплаты. Сначала отмените/скорректируйте оплаты.', http_status=409)
        try:
            validate_sale_transition(sale.sale_status, Sale.STATUS_CANCELED)
        except ValueError as e:
            return _err('INVALID_TRANSITION', str(e), http_status=422)
        with transaction.atomic():
            try:
                reverse_warehouse_for_sale(sale)
            except Exception as e:
                return _err('WAREHOUSE_ROLLBACK', str(e), http_status=422)
            restore_reservations_for_sale(sale=sale, user=request.user, request=request)
            sale.sale_status = Sale.STATUS_CANCELED
            sale.save(update_fields=['sale_status', 'updated_at'])
        if sale.linked_order_id:
            from .order_sync import recalculate_order_line_shipped_from_sale_lines_for_order
            recalculate_order_line_shipped_from_sale_lines_for_order(sale.linked_order_id)
        from .commercial_audit import log_commercial_audit
        log_commercial_audit(
            user=request.user,
            request=request,
            section='sales',
            description=f'Отмена продажи #{sale.order_number}',
            model_cls=Sale,
            instance=sale,
            payload_extra={'action': 'cancel_sale', 'sale_id': sale.pk},
        )
        return Response(SaleSerializer(sale, context={'request': request}).data)

    @action(detail=True, methods=['get'], url_path='waybill')
    def waybill(self, request, pk=None):
        return self._serve_nakladnaya(request)

    @action(detail=True, methods=['get'], url_path='credit-check')
    def credit_check_for_sale(self, request, pk=None):
        """Проверить кредитный лимит клиента перед оплатой/отгрузкой по продаже."""
        from .credit_check import check_credit_limit, credit_check_result_to_dict
        sale = self.get_object()
        if not sale.client_id:
            return _err('NO_CLIENT', 'У продажи не указан клиент')
        result = check_credit_limit(sale.client, additional_amount=sale.revenue or Decimal('0'))
        return Response(credit_check_result_to_dict(result))

    @action(detail=True, methods=['patch'], url_path='status')
    def set_status(self, request, pk=None):
        """
        Изменить статус продажи через централизованный state machine.

        Проверки перед shipped/closed:
          - Кредитный лимит (hard)
          - Наличие склада (если продажа привязана к партии)
        Допускает force_credit_override=true для пользователей с правом credit_limit_override.
        """
        from .state_machine import validate_sale_transition, validate_sale_ship
        from .credit_check import enforce_credit_limit, CreditLimitBlocked

        sale = self.get_object()
        new_status = request.data.get('status')
        if not new_status:
            return _err('MISSING_STATUS', 'Укажите status')

        try:
            validate_sale_transition(sale.sale_status, new_status)
        except ValueError as e:
            return _err('INVALID_STATUS_TRANSITION', str(e), http_status=422)

        shipping_statuses = (Sale.STATUS_SHIPPED, Sale.STATUS_CLOSED, Sale.STATUS_PARTIALLY_SHIPPED)
        if new_status in shipping_statuses:
            # Stock / reservation check
            try:
                validate_sale_ship(sale)
            except ValueError as e:
                return _err('SHIP_BLOCKED', str(e), http_status=422)

            # Hard credit limit check
            if sale.client_id:
                force_override = str(request.data.get('force_credit_override', '')).lower() in ('1', 'true', 'yes')
                if force_override:
                    sale.credit_limit_bypassed = True
                    sale.save(update_fields=['credit_limit_bypassed'])
                try:
                    enforce_credit_limit(
                        sale.client,
                        Decimal('0'),
                        user=request.user,
                        force_override=force_override,
                    )
                except CreditLimitBlocked as e:
                    return _err('CREDIT_LIMIT_BLOCKED', str(e), http_status=422)

        sale.sale_status = new_status
        sale.save(update_fields=['sale_status', 'updated_at'])
        sale.refresh_from_db()
        from .sale_warehouse import apply_warehouse_for_sale
        from .reservations import auto_fulfill_sale_lines_after_shipping
        try:
            apply_warehouse_for_sale(sale)
        except (ValueError, Exception) as e:
            from rest_framework.exceptions import ValidationError as DrfV
            if isinstance(e, DrfV):
                return _err('WAREHOUSE_APPLY', str(e.detail), http_status=422)
            return _err('WAREHOUSE_APPLY', str(e), http_status=422)
        if sale.linked_order_id and new_status in shipping_statuses:
            auto_fulfill_sale_lines_after_shipping(
                sale=sale,
                order=sale.linked_order,
                user=request.user if request.user.is_authenticated else None,
                request=request,
            )
        return Response(SaleSerializer(sale, context={'request': request}).data)

    @action(detail=True, methods=['get'], url_path='receipt')
    def receipt(self, request, pk=None):
        """HTML-квитанция об оплате."""
        sale = self.get_object()
        payments = list(sale.payments.filter(status=Payment.STATUS_ACTIVE).all())
        parts = [
            '<!DOCTYPE html><html lang="ru"><head><meta charset="utf-8">',
            f'<title>Квитанция {escape(sale.receipt_number or sale.order_number)}</title>',
            '<style>body{font-family:system-ui,sans-serif;max-width:600px;margin:2rem auto;line-height:1.5;}</style>',
            '</head><body><h1>Квитанция об оплате</h1>',
            f'<p><strong>Продажа №</strong> {escape(sale.order_number)} от {sale.date.isoformat()}</p>',
        ]
        if sale.client_id:
            parts.append(f'<p><strong>Клиент:</strong> {escape(sale.client.name)}</p>')
        if payments:
            for p in payments:
                parts.append(
                    f'<p>{p.date.isoformat()} — {p.get_payment_type_display()}: '
                    f'<strong>{p.amount}</strong> ({p.get_payment_method_display()})</p>'
                )
        else:
            parts.append('<p>Оплаты не зафиксированы</p>')
        parts.append('<p><em>Сформировано автоматически.</em></p></body></html>')
        html = ''.join(parts)
        resp = HttpResponse(html, content_type='text/html; charset=utf-8')
        resp['Content-Disposition'] = f'inline; filename="sale-receipt-{sale.id}.html"'
        return resp


# ─────────────────────────────────────────────────────────────────────────────
# PAYMENT (Оплата)
# ─────────────────────────────────────────────────────────────────────────────

class PaymentViewSet(ActivityLoggingMixin, viewsets.ModelViewSet):
    queryset = Payment.objects.select_related('client', 'linked_order', 'linked_sale', 'linked_return', 'created_by').all()
    serializer_class = PaymentSerializer
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'payments'
    activity_section = 'Оплаты'
    activity_label = 'оплата'
    ordering_fields = ['id', 'date']
    filterset_class = PaymentFilter

    def destroy(self, request, *args, **kwargs):
        return Response(
            {'code': 'DELETE_DISABLED', 'error': 'Удаление оплат отключено. Используйте /api/payments/{id}/cancel/.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    @action(detail=True, methods=['post', 'patch'], url_path='cancel')
    def cancel_payment(self, request, pk=None):
        p = self.get_object()
        if p.status == Payment.STATUS_CANCELED:
            return _err('PAYMENT_ALREADY_CANCELED', 'Оплата уже отменена', http_status=422)
        p.status = Payment.STATUS_CANCELED
        p.save(update_fields=['status'])
        from .commercial_audit import log_commercial_audit
        log_commercial_audit(
            user=request.user,
            request=request,
            section='payments',
            description=f'Отмена оплаты #{p.payment_number or p.pk}',
            model_cls=Payment,
            instance=p,
            payload_extra={'action': 'cancel_payment', 'payment_id': p.pk},
        )
        return Response(PaymentSerializer(p).data)

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(created_by=user)
        inst = serializer.instance
        from apps.realtime.broadcast import schedule_push
        from django.db import transaction
        transaction.on_commit(lambda: schedule_push(
            resource='payment',
            action='created',
            entity_id=inst.pk,
            extra={'client_id': inst.client_id, 'payment_type': inst.payment_type},
        ))

    @action(detail=False, methods=['get'], url_path='select-sources')
    def select_sources(self, request):
        client_id = request.query_params.get('client_id')
        sale_id = request.query_params.get('sale_id')
        order_id = request.query_params.get('order_id')
        return_id = request.query_params.get('return_id')

        from .payment_status import order_payment_metrics, sale_payment_metrics

        clients_qs = Client.objects.filter(is_active=True).order_by('name', 'id')
        if client_id:
            clients_qs = clients_qs.filter(pk=client_id)
        clients = [{'id': c.id, 'label': c.name} for c in clients_qs[:200]]

        orders_qs = Order.objects.select_related('client').order_by('-date', '-id')
        sales_qs = Sale.objects.select_related('client').order_by('-date', '-id')
        returns_qs = Return.objects.select_related('sale', 'sale__client').filter(status=Return.STATUS_COMPLETED).order_by('-date', '-id')

        if client_id:
            orders_qs = orders_qs.filter(client_id=client_id)
            sales_qs = sales_qs.filter(client_id=client_id)
            returns_qs = returns_qs.filter(sale__client_id=client_id)
        if order_id:
            orders_qs = orders_qs.filter(pk=order_id)
        if sale_id:
            sales_qs = sales_qs.filter(pk=sale_id)
            returns_qs = returns_qs.filter(sale_id=sale_id)
        if return_id:
            returns_qs = returns_qs.filter(pk=return_id)

        orders_payload = []
        for order in orders_qs[:200]:
            m = order_payment_metrics(order)
            debt = api_decimal_str(m['debt_amount'])
            orders_payload.append(
                {
                    'id': order.id,
                    'label': f'{order.order_number} — долг {debt}',
                    'client': order.client_id,
                    'debt_amount': debt,
                    'payment_status': m['payment_status'],
                    'status': order.status,
                },
            )

        sales_payload = []
        for sale in sales_qs[:200]:
            m = sale_payment_metrics(sale)
            debt_amount_raw = Decimal(str(m['debt_amount'] or 0))
            if debt_amount_raw <= 0:
                continue
            debt = api_decimal_str(debt_amount_raw)
            sales_payload.append(
                {
                    'id': sale.id,
                    'client_id': sale.client_id,
                    'label': f'{sale.sale_number or sale.order_number} — долг {debt}',
                    'client': sale.client_id,
                    'debt_amount': debt,
                    'payment_status': m['payment_status'],
                    'sale_status': sale.sale_status,
                },
            )

        returns_payload = []
        for ret in returns_qs[:200]:
            amount = Decimal('0')
            for line in ret.lines.select_related('sale_line').all():
                unit_price = Decimal(str(line.sale_line.unit_price or 0)) if line.sale_line_id else Decimal('0')
                amount += Decimal(str(line.quantity or 0)) * unit_price
            amount_str = api_decimal_str(amount)
            returns_payload.append(
                {
                    'id': ret.id,
                    'label': f'{ret.return_number or f"RET-{ret.id}"} — к возврату {amount_str}',
                    'client': ret.sale.client_id if ret.sale_id else None,
                    'return_amount': amount_str,
                    'status': ret.status,
                },
            )

        return Response(
            {
                'clients': clients,
                'orders': orders_payload,
                'sales': sales_payload,
                'returns': returns_payload,
            },
        )

    @action(detail=False, methods=['get'], url_path='summary')
    def summary(self, request):
        """Сводка оплат по клиенту."""
        client_id = request.query_params.get('client_id')
        if not client_id:
            return _err('MISSING_CLIENT', 'Укажите client_id')
        try:
            client = Client.objects.get(pk=client_id)
        except Client.DoesNotExist:
            return _err('NOT_FOUND', 'Клиент не найден', http_status=404)

        payments = Payment.objects.filter(client=client, status=Payment.STATUS_ACTIVE)
        sales = Sale.objects.filter(client=client).exclude(sale_status__in=(Sale.STATUS_DRAFT, Sale.STATUS_CANCELED))

        total_paid = payments.filter(
            payment_type__in=[Payment.TYPE_PREPAYMENT, Payment.TYPE_PAYMENT, Payment.TYPE_SURCHARGE]
        ).aggregate(s=Sum('amount'))['s'] or Decimal('0')
        total_refunded = payments.filter(
            payment_type=Payment.TYPE_REFUND
        ).aggregate(s=Sum('amount'))['s'] or Decimal('0')
        net_paid = total_paid - total_refunded
        total_revenue = sales.aggregate(s=Sum('revenue'))['s'] or Decimal('0')

        from config.api_numbers import api_decimal_str
        return Response({
            'client_id': client.id,
            'client_name': client.name,
            'total_paid_gross': api_decimal_str(total_paid),
            'total_refunded': api_decimal_str(total_refunded),
            'total_paid_net': api_decimal_str(net_paid),
            'total_revenue': api_decimal_str(total_revenue),
            'client_debt_money': api_decimal_str(max(Decimal('0'), total_revenue - net_paid)),
            'client_advance_amount': api_decimal_str(max(Decimal('0'), net_paid - total_revenue)),
        })


# ─────────────────────────────────────────────────────────────────────────────
# RETURN (Возврат)
# ─────────────────────────────────────────────────────────────────────────────

class ReturnViewSet(ActivityLoggingMixin, viewsets.ModelViewSet):
    queryset = Return.objects.select_related(
        'sale', 'sale__client', 'linked_order', 'created_by',
    ).prefetch_related('lines').all()
    serializer_class = ReturnSerializer
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'returns'
    activity_section = 'Возвраты'
    activity_label = 'возврат'
    ordering_fields = ['id', 'date']
    filterset_class = ReturnFilter

    def _recalculate_sale_after_return(self, sale: Sale) -> None:
        sale_lines = list(SaleLine.objects.filter(sale=sale).values('id', 'quantity'))
        if not sale_lines:
            return
        sold_total = sum((Decimal(str(sl['quantity'] or 0)) for sl in sale_lines), Decimal('0'))
        returned_total = ReturnLine.objects.filter(
            sale_line_id__in=[sl['id'] for sl in sale_lines],
            return_doc__status=Return.STATUS_COMPLETED,
        ).aggregate(s=Sum('quantity'))['s'] or Decimal('0')
        if sold_total > 0 and Decimal(str(returned_total)) >= sold_total:
            if sale.sale_status != Sale.STATUS_CANCELED:
                sale.sale_status = Sale.STATUS_CANCELED
                sale.save(update_fields=['sale_status', 'updated_at'])
        elif sale.sale_status == Sale.STATUS_DRAFT and sale.warehouse_stock_applied:
            sale.sale_status = Sale.STATUS_SHIPPED
            sale.save(update_fields=['sale_status', 'updated_at'])
        elif sale.sale_status == Sale.STATUS_CANCELED:
            sale.sale_status = Sale.STATUS_SHIPPED
            sale.save(update_fields=['sale_status', 'updated_at'])

    def _next_payment_number(self, payment_date) -> str:
        year = (payment_date or timezone.now().date()).year
        last = Payment.objects.filter(payment_number__startswith=f'PAY-{year}-').order_by('-payment_number').first()
        try:
            last_n = int(last.payment_number.split('-')[-1]) if last else 0
        except (ValueError, IndexError):
            last_n = 0
        return f'PAY-{year}-{last_n + 1:04d}'

    def _create_auto_refund_payment(self, ret_doc: Return, user) -> Payment | None:
        sale = ret_doc.sale
        total_return_amount = return_document_amount(ret_doc)
        if total_return_amount <= 0:
            return None
        active_payments = Payment.objects.filter(status=Payment.STATUS_ACTIVE).filter(
            Q(linked_sale=sale) | Q(linked_return__sale=sale),
        ).distinct()
        incoming = sum(
            (Decimal(str(p.amount or 0)) for p in active_payments if p.payment_type in (
                Payment.TYPE_PREPAYMENT,
                Payment.TYPE_PAYMENT,
                Payment.TYPE_SURCHARGE,
            )),
            Decimal('0'),
        )
        existing_refunds = sum(
            (Decimal(str(p.amount or 0)) for p in active_payments if p.payment_type == Payment.TYPE_REFUND),
            Decimal('0'),
        )
        order_applied = Decimal(str(sale.order_paid_amount_applied or 0))
        net_paid = incoming + order_applied - existing_refunds
        active_due = sale_active_total_amount(sale)
        refundable = min(total_return_amount, max(Decimal('0'), net_paid - active_due)).quantize(Decimal('0.01'))
        if refundable <= 0:
            return None
        if Payment.objects.filter(
            linked_return=ret_doc,
            payment_type=Payment.TYPE_REFUND,
            status=Payment.STATUS_ACTIVE,
        ).exists():
            return None
        return Payment.objects.create(
            payment_number=self._next_payment_number(ret_doc.date),
            date=ret_doc.date,
            client=sale.client,
            linked_order=ret_doc.linked_order,
            linked_sale=sale,
            linked_return=ret_doc,
            payment_type=Payment.TYPE_REFUND,
            payment_method=Payment.METHOD_CASH,
            amount=refundable,
            status=Payment.STATUS_ACTIVE,
            comment=f'Автоматический refund по возврату #{ret_doc.return_number or ret_doc.pk}',
            created_by=user if getattr(user, 'is_authenticated', False) else None,
        )

    def destroy(self, request, *args, **kwargs):
        return Response(
            {'code': 'DELETE_DISABLED', 'error': 'Удаление возвратов отключено. Используйте /api/returns/{id}/cancel/.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    @action(detail=True, methods=['post', 'patch'], url_path='cancel')
    def cancel_return(self, request, pk=None):
        from .return_rollback import rollback_return_document

        ret_doc = self.get_object()
        if ret_doc.status == Return.STATUS_CANCELED:
            return _err('RETURN_ALREADY_CANCELED', 'Возврат уже отменён', http_status=422)
        with transaction.atomic():
            if ret_doc.status == Return.STATUS_COMPLETED:
                from apps.warehouse.models import WarehouseBatch as WarehouseBatchModel

                for line in ret_doc.lines.select_related('rework_receipt_batch').all():
                    if line.return_target != ReturnLine.TARGET_REWORK:
                        continue
                    wb_id = line.rework_receipt_batch_id
                    if not wb_id:
                        continue
                    if SaleLine.objects.filter(warehouse_batch_id=wb_id).exists():
                        return _err(
                            'DOWNSTREAM_USED',
                            'Нельзя отменить возврат: партия переделанных уже связана с продажей.',
                            http_status=409,
                        )
                    wb_row = WarehouseBatchModel.objects.filter(pk=wb_id).first()
                    if wb_row is not None and wb_row.status != WarehouseBatchModel.STATUS_AVAILABLE:
                        return _err(
                            'DOWNSTREAM_USED',
                            'Нельзя отменить возврат: партия переделанных уже использована на складе.',
                            http_status=409,
                        )

                line_ids = list(ret_doc.lines.values_list('id', flat=True))
                defects = list(
                    DefectRecord.objects.filter(
                        source_type=DefectRecord.SOURCE_RETURN,
                        source_id__in=line_ids,
                    ),
                )
                for d in defects:
                    used_statuses = {
                        DefectRecord.STATUS_SOLD,
                        DefectRecord.STATUS_WRITTEN_OFF,
                        DefectRecord.STATUS_SENT_TO_REWORK,
                        DefectRecord.STATUS_REWORKED,
                        DefectRecord.STATUS_CLOSED,
                    }
                    counters_changed = any(
                        Decimal(str(x or 0)) > 0
                        for x in (
                            d.sold_quantity_pcs,
                            d.written_off_quantity_pcs,
                            d.sent_to_rework_quantity_pcs,
                        )
                    )
                    if d.status in used_statuses or counters_changed:
                        return _err(
                            'DOWNSTREAM_USED',
                            'Нельзя отменить возврат: связанный defect уже использован.',
                            http_status=409,
                        )
                reworks = ReworkRequest.objects.filter(defect_record_id__in=[d.id for d in defects])
                for rw in reworks:
                    if rw.status in (
                        ReworkRequest.STATUS_IN_PROGRESS,
                        ReworkRequest.STATUS_COMPLETED,
                        ReworkRequest.STATUS_CANCELED,
                    ):
                        return _err(
                            'DOWNSTREAM_USED',
                            'Нельзя отменить возврат: связанная переделка уже использована.',
                            http_status=409,
                        )

                for line in ret_doc.lines.select_related('sale_line', 'sale_line__warehouse_batch').all():
                    if line.return_target != ReturnLine.TARGET_WAREHOUSE:
                        continue
                    wb = None
                    if line.sale_line_id and line.sale_line.warehouse_batch_id:
                        wb = line.sale_line.warehouse_batch
                    elif ret_doc.sale and ret_doc.sale.warehouse_batch_id:
                        wb = ret_doc.sale.warehouse_batch
                    if wb is None:
                        continue
                    current = Decimal(str(wb.quantity or 0))
                    delta = Decimal(str(line.quantity or 0))
                    if current - delta < 0:
                        return _err(
                            'WAREHOUSE_ROLLBACK_NEGATIVE',
                            'Нельзя откатить возврат: rollback уведет складской остаток в минус.',
                            http_status=409,
                        )
                try:
                    rollback_return_document(ret_doc)
                except Exception as exc:
                    return _err(
                        'RETURN_ROLLBACK_FAILED',
                        f'Не удалось откатить возврат: {exc}',
                        http_status=409,
                    )
                Payment.objects.filter(
                    linked_return=ret_doc,
                    payment_type=Payment.TYPE_REFUND,
                    status=Payment.STATUS_ACTIVE,
                ).update(status=Payment.STATUS_CANCELED)
            ret_doc.status = Return.STATUS_CANCELED
            ret_doc.save(update_fields=['status'])
            self._recalculate_sale_after_return(ret_doc.sale)
        from .commercial_audit import log_commercial_audit
        log_commercial_audit(
            user=request.user,
            request=request,
            section='returns',
            description=f'Отмена возврата #{ret_doc.return_number or ret_doc.pk}',
            model_cls=Return,
            instance=ret_doc,
            payload_extra={'action': 'cancel_return', 'return_id': ret_doc.pk},
        )
        return Response(ReturnSerializer(ret_doc, context={'request': request}).data)

    @action(detail=True, methods=['post', 'patch'], url_path='complete')
    def complete_return(self, request, pk=None):
        """Провести возврат: склад/брак/переделка только здесь (из статуса draft)."""
        ret_doc = self.get_object()
        if ret_doc.status == Return.STATUS_COMPLETED:
            return _err('RETURN_ALREADY_COMPLETED', 'Возврат уже проведен.', http_status=422)
        if ret_doc.status == Return.STATUS_CANCELED:
            return _err('RETURN_ALREADY_CANCELED', 'Отмененный возврат нельзя провести.', http_status=422)
        if ret_doc.status != Return.STATUS_DRAFT:
            return _err('RETURN_COMPLETE_FAILED', 'Провести можно только возврат в статусе draft.', http_status=422)
        if not ret_doc.lines.exists():
            return _err('NO_LINES', 'Нет строк возврата', http_status=422)
        ser = ReturnSerializer()
        with transaction.atomic():
            try:
                ser.apply_completion_effects(ret_doc)
            except ValidationError as exc:
                detail = getattr(exc, 'detail', {})
                code = 'RETURN_COMPLETE_FAILED'
                message = 'Ошибка проведения возврата.'
                if isinstance(detail, dict):
                    code = str(detail.get('code') or code)
                    message = str(detail.get('detail') or detail.get('error') or message)
                return _err(code, message, http_status=400)
            except Exception as exc:
                return _err('RETURN_COMPLETE_FAILED', f'Ошибка проведения возврата: {exc}', http_status=409)
            ret_doc.status = Return.STATUS_COMPLETED
            ret_doc.save(update_fields=['status'])
            self._recalculate_sale_after_return(ret_doc.sale)
            self._create_auto_refund_payment(ret_doc, request.user)
        return Response(ReturnSerializer(ret_doc, context={'request': request}).data)

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(created_by=user)
        inst = serializer.instance
        from apps.realtime.broadcast import schedule_push
        from django.db import transaction
        transaction.on_commit(lambda: schedule_push(
            resource='return',
            action='created',
            entity_id=inst.pk,
            extra={'sale_id': inst.sale_id},
        ))

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        data = ReturnSerializer(instance, context={'request': request}).data
        downstream = []
        line_ids = list(instance.lines.values_list('id', flat=True))
        warehouse_effects = []
        for line in instance.lines.select_related('sale_line', 'sale_line__warehouse_batch').all():
            if line.return_target != ReturnLine.TARGET_WAREHOUSE:
                continue
            batch_id = None
            if line.sale_line_id and line.sale_line.warehouse_batch_id:
                batch_id = line.sale_line.warehouse_batch_id
            elif instance.sale and instance.sale.warehouse_batch_id:
                batch_id = instance.sale.warehouse_batch_id
            warehouse_effects.append(
                {
                    'type': 'warehouse_effect',
                    'source_return_line_id': line.id,
                    'warehouse_batch_id': batch_id,
                    'quantity': api_decimal_str(Decimal(str(line.quantity or 0))),
                },
            )
        defects = DefectRecord.objects.filter(
            source_type=DefectRecord.SOURCE_RETURN,
            source_id__in=line_ids,
        ).values('id', 'status', 'source_id')
        defect_ids = [d['id'] for d in defects]
        reworks = ReworkRequest.objects.filter(defect_record_id__in=defect_ids).values(
            'id', 'status', 'defect_record_id', 'result_warehouse_batch_id',
        )
        data['linked_entities'] = {
            'sale': {
                'id': instance.sale_id,
                'label': instance.sale.order_number,
            },
            'client': (
                {
                    'id': instance.sale.client_id,
                    'label': instance.sale.client.name,
                } if instance.sale and instance.sale.client_id else None
            ),
            'linked_order': (
                {
                    'id': instance.linked_order_id,
                    'label': instance.linked_order.order_number,
                } if instance.linked_order_id else None
            ),
        }
        for d in defects:
            downstream.append({
                'type': 'defect_record',
                'id': d['id'],
                'status': d['status'],
                'source_return_line_id': d['source_id'],
            })
        for r in reworks:
            downstream.append({
                'type': 'rework_request',
                'id': r['id'],
                'status': r['status'],
                'defect_record_id': r['defect_record_id'],
                'result_warehouse_batch_id': r['result_warehouse_batch_id'],
            })
        for line in instance.lines.select_related('rework_receipt_batch').all():
            if line.rework_receipt_batch_id:
                downstream.append({
                    'type': 'warehouse_batch',
                    'id': line.rework_receipt_batch_id,
                    'warehouse_batch_id': line.rework_receipt_batch_id,
                    'stock_bucket': 'reworked',
                    'label': 'Переделанные (склад)',
                    'source_return_line_id': line.id,
                })
        active_refunds = Payment.objects.filter(
            linked_return=instance,
            payment_type=Payment.TYPE_REFUND,
            status=Payment.STATUS_ACTIVE,
        ).values('id', 'payment_number', 'amount', 'status')
        for p in active_refunds:
            downstream.append(
                {
                    'type': 'payment',
                    'id': p['id'],
                    'label': f"Возврат денег №{p['payment_number'] or p['id']} — {api_decimal_str(Decimal(str(p['amount'] or 0)))} сом",
                    'payment_number': p['payment_number'],
                    'amount': api_decimal_str(Decimal(str(p['amount'] or 0))),
                    'status': p['status'],
                },
            )
        downstream.extend(warehouse_effects)
        data['downstream_links'] = downstream
        data['available_status_transitions'] = []
        data['available_actions'] = {
            'waybill': True,
            'complete': instance.status == Return.STATUS_DRAFT,
        }
        return Response(data)

    @action(detail=False, methods=['get'], url_path='select-sources')
    def select_sources(self, request):
        sale_id = request.query_params.get('sale_id')
        q = (request.query_params.get('q') or '').strip().lower()
        try:
            page = max(1, int(request.query_params.get('page') or 1))
        except (TypeError, ValueError):
            page = 1
        try:
            page_size = int(request.query_params.get('page_size') or 20)
        except (TypeError, ValueError):
            page_size = 20
        page_size = min(max(page_size, 1), 100)

        sales_qs = Sale.objects.select_related('client').exclude(
            sale_status=Sale.STATUS_CANCELED,
        ).order_by('-date', '-id')
        all_sales_payload = []
        for sale in sales_qs:
            if not sale_return_source_is_valid(sale):
                continue
            sale_lines = (
                SaleLine.objects.filter(sale=sale)
                .select_related('warehouse_batch__profile')
                .values(
                    'id',
                    'product',
                    'quantity',
                    'unit_price',
                    'warehouse_batch__profile__name',
                )
            )
            total_returnable = Decimal('0')
            product_names = []
            for sl in sale_lines:
                product_name = (
                    (sl.get('product') or '').strip()
                    or (sl.get('warehouse_batch__profile__name') or '').strip()
                )
                if product_name and product_name not in product_names:
                    product_names.append(product_name)
                sold = Decimal(str(sl['quantity'] or 0))
                returned = ReturnLine.objects.filter(
                    sale_line_id=sl['id'],
                    return_doc__status=Return.STATUS_COMPLETED,
                ).aggregate(s=Sum('quantity'))['s'] or Decimal('0')
                returnable = sold - Decimal(str(returned))
                if returnable > 0:
                    total_returnable += returnable
            if total_returnable <= 0:
                continue
            sale_sum = api_decimal_str(Decimal(str(sale.revenue or 0)))
            client_name = sale.client.name if sale.client_id else ''
            sale_date = sale.date.isoformat() if sale.date else ''
            products = ', '.join(product_names)
            sale_number = sale.sale_number or str(sale.id)
            haystack = ' '.join(
                str(x)
                for x in (
                    sale.id,
                    sale.sale_number,
                    sale.order_number,
                    client_name,
                    sale_date,
                    sale_sum,
                    products,
                )
                if x not in (None, '')
            ).lower()
            if q and q not in haystack:
                continue
            all_sales_payload.append({
                'id': sale.id,
                'sale_number': sale_number,
                'date': sale_date,
                'client': sale.client_id,
                'client_name': client_name,
                'total_amount': sale_sum,
                'products': products,
                'label': f'Продажа №{sale_number} — {sale_date} — {client_name or "—"} — {sale_sum} сом — {products or "—"}',
                'status': 'completed',
                'sale_status': sale.sale_status,
                'returnable_quantity': api_decimal_str(total_returnable),
            })

        total = len(all_sales_payload)
        pages = (total + page_size - 1) // page_size
        start = (page - 1) * page_size
        end = start + page_size
        sales_payload = all_sales_payload[start:end]
        lines = []
        if sale_id:
            sale_lines = (
                SaleLine.objects.filter(sale_id=sale_id)
                .order_by('id')
                .values('id', 'product', 'quantity', 'unit_price')
            )
            for sl in sale_lines:
                sold = Decimal(str(sl['quantity'] or 0))
                returned = ReturnLine.objects.filter(
                    sale_line_id=sl['id'],
                    return_doc__status=Return.STATUS_COMPLETED,
                ).aggregate(s=Sum('quantity'))['s'] or Decimal('0')
                returnable = sold - Decimal(str(returned))
                if returnable <= 0:
                    continue
                sold_s = api_decimal_str(sold)
                ret_s = api_decimal_str(Decimal(str(returned)))
                retable_s = api_decimal_str(returnable)
                unit_price_s = api_decimal_str(Decimal(str(sl['unit_price'] or 0)))
                lines.append(
                    {
                        'id': sl['id'],
                        'label': f"{sl['product']} — продано {sold_s} — доступно {retable_s}",
                        'product': sl['product'],
                        'product_name': sl['product'],
                        'quantity': sold_s,
                        'quantity_sold': sold_s,
                        'sold_quantity': sold_s,
                        'returned_quantity': ret_s,
                        'returnable_quantity': retable_s,
                        'unit_price': unit_price_s,
                    },
                )
        return Response({
            'sales': sales_payload,
            'sale_lines': lines,
            'meta': {
                'page': page,
                'page_size': page_size,
                'total': total,
                'pages': pages,
            },
        })

    @action(detail=True, methods=['get'], url_path='waybill')
    def waybill(self, request, pk=None):
        """HTML-документ возврата (канонический endpoint)."""
        ret_doc = self.get_object()
        return _return_html_response(ret_doc)


def _return_html_response(ret_doc: Return) -> HttpResponse:
    parts = [
        '<!DOCTYPE html><html lang="ru"><head><meta charset="utf-8">',
        f'<title>Возврат {escape(ret_doc.return_number or str(ret_doc.id))}</title>',
        '<style>body{font-family:system-ui,sans-serif;max-width:720px;margin:2rem auto;line-height:1.45;}',
        'table{border-collapse:collapse;width:100%;}td,th{border:1px solid #ccc;padding:0.4rem 0.6rem;text-align:left;}',
        'th{background:#f5f5f5;}</style>',
        '</head><body>',
        '<h1>Акт возврата товара</h1>',
        f'<p><strong>№</strong> {escape(ret_doc.return_number or str(ret_doc.id))} '
        f'<strong>от</strong> {ret_doc.date.isoformat()}</p>',
        f'<p><strong>К продаже №:</strong> {escape(ret_doc.sale.order_number)}</p>',
    ]
    if ret_doc.sale.client_id:
        parts.append(f'<p><strong>Клиент:</strong> {escape(ret_doc.sale.client.name)}</p>')
    if ret_doc.return_reason:
        parts.append(f'<p><strong>Причина возврата:</strong> {escape(ret_doc.return_reason)}</p>')
    parts.append(
        '<table><thead><tr>'
        '<th>№</th><th>Товар</th><th>Кол-во</th><th>Состояние</th><th>Назначение</th><th>Комментарий</th>'
        '</tr></thead><tbody>'
    )
    for i, line in enumerate(ret_doc.lines.all(), 1):
        parts.append(
            f'<tr>'
            f'<td>{i}</td>'
            f'<td>{escape(line.product or "")}</td>'
            f'<td>{escape(str(line.quantity))}</td>'
            f'<td>{escape(line.get_condition_type_display())}</td>'
            f'<td>{escape(line.get_return_target_display())}</td>'
            f'<td>{escape(line.comment or "")}</td>'
            f'</tr>'
        )
    parts.append('</tbody></table>')
    if ret_doc.comment:
        parts.append(f'<p><strong>Комментарий:</strong> {escape(ret_doc.comment)}</p>')
    parts.append('<p><em>Сформировано автоматически.</em></p></body></html>')
    html = ''.join(parts)
    resp = HttpResponse(html, content_type='text/html; charset=utf-8')
    resp['Content-Disposition'] = f'inline; filename="return-waybill-{ret_doc.id}.html"'
    return resp


# ─────────────────────────────────────────────────────────────────────────────
# DEFECT RECORD (Брак)
# ─────────────────────────────────────────────────────────────────────────────

class DefectRecordViewSet(ActivityLoggingMixin, viewsets.ModelViewSet):
    queryset = DefectRecord.objects.select_related('profile', 'created_by').all()
    serializer_class = DefectRecordSerializer
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'defects'
    activity_section = 'Брак'
    activity_label = 'запись брака'
    ordering_fields = ['id', 'created_at', 'status']
    filterset_class = DefectFilter

    def destroy(self, request, *args, **kwargs):
        return Response(
            {'code': 'DELETE_DISABLED', 'error': 'Удаление записей брака отключено (используйте writeoff / cancel сценарии).'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(created_by=user)
        inst = serializer.instance
        from apps.realtime.broadcast import schedule_push
        from django.db import transaction
        transaction.on_commit(lambda: schedule_push(
            resource='defect_record',
            action='created',
            entity_id=inst.pk,
            extra={'source_type': inst.source_type, 'status': inst.status},
        ))

    def perform_update(self, serializer):
        super().perform_update(serializer)
        inst = serializer.instance
        from apps.realtime.broadcast import schedule_push
        from django.db import transaction
        transaction.on_commit(lambda: schedule_push(
            resource='defect_record',
            action='updated',
            entity_id=inst.pk,
            extra={'status': inst.status},
        ))

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        data = DefectRecordSerializer(instance, context={'request': request}).data
        from .state_machine import DEFECT_TRANSITIONS
        linked_source = None
        if instance.source_type == DefectRecord.SOURCE_RETURN and instance.source_id:
            rl = ReturnLine.objects.select_related('return_doc', 'sale_line').filter(pk=instance.source_id).first()
            if rl is not None:
                linked_source = {
                    'return_line_id': rl.pk,
                    'return_doc_id': rl.return_doc_id,
                    'sale_line_id': rl.sale_line_id,
                    'label': f"ReturnLine #{rl.pk}: {rl.product} × {rl.quantity}",
                }
        reworks = list(
            ReworkRequest.objects.filter(defect_record=instance).values(
                'id', 'status', 'result_warehouse_batch_id', 'original_sale_id',
            )
        )
        data['linked_entities'] = {
            'source': linked_source,
            'rework_requests': reworks,
        }
        data['available_status_transitions'] = DEFECT_TRANSITIONS.get(instance.status, [])
        rem = Decimal(str(instance.quantity_pcs or 0))
        has_remainder = rem > Decimal('0.0001')
        data['available_actions'] = {
            'send_to_rework': has_remainder and DefectRecord.STATUS_SENT_TO_REWORK in DEFECT_TRANSITIONS.get(instance.status, []),
            'complete_rework': False,
            'writeoff': has_remainder and DefectRecord.STATUS_WRITTEN_OFF in DEFECT_TRANSITIONS.get(instance.status, []),
            'sell': has_remainder and instance.status in (
                DefectRecord.STATUS_NEW,
                DefectRecord.STATUS_ON_STOCK,
                DefectRecord.STATUS_REWORKED,
            ),
        }
        return Response(data)

    @action(detail=False, methods=['get'], url_path='select-sources')
    def select_sources(self, request):
        from apps.warehouse.models import WarehouseBatch

        lines = (
            ReturnLine.objects.select_related('return_doc')
            .exclude(
                id__in=DefectRecord.objects.filter(source_type=DefectRecord.SOURCE_RETURN).values_list('source_id', flat=True),
            )
            .order_by('-id')
            .values('id', 'product', 'quantity', 'return_doc_id', 'return_doc__return_number', 'condition_type')[:300]
        )
        wh_def = list(
            WarehouseBatch.objects.filter(
                quality=WarehouseBatch.QUALITY_DEFECT,
                status=WarehouseBatch.STATUS_AVAILABLE,
                quantity__gt=0,
                linked_defect_record__isnull=True,
            )
            .order_by('-date', '-id')
            .values('id', 'product', 'quantity', 'quality', 'status')[:200]
        )
        return Response({
            'return_lines': [
                {
                    'id': rl['id'],
                    'label': (
                        f"{rl['product']} — возврат {api_decimal_str(Decimal(str(rl['quantity'] or 0)))} шт — "
                        f"причина: {rl['condition_type']}"
                    ),
                    'product': rl['product'],
                    'quantity_pcs': api_decimal_str(Decimal(str(rl['quantity'] or 0))),
                    'return_id': rl['return_doc_id'],
                    'return_number': rl['return_doc__return_number'] or f"RET-{rl['return_doc_id']}",
                }
                for rl in lines
            ],
            'warehouse_defect_batches': [
                {
                    'id': b['id'],
                    'label': f"#{b['id']} — {b['product']} — доступно {api_decimal_str(Decimal(str(b['quantity'] or 0)))} шт — Брак",
                    'product': b['product'],
                    'available_quantity_pcs': api_decimal_str(Decimal(str(b['quantity'] or 0))),
                    'quantity_pcs': api_decimal_str(Decimal(str(b['quantity'] or 0))),
                    'quality': b['quality'],
                    'status': b['status'],
                }
                for b in wh_def
            ],
        })

    @action(detail=True, methods=['post'], url_path='send-to-rework')
    def send_to_rework(self, request, pk=None):
        """Создать ReworkRequest; опционально часть остатка (body quantity)."""
        from django.utils import timezone as tz

        from .state_machine import validate_defect_transition

        record = self.get_object()
        if ReworkRequest.objects.filter(
            defect_record=record,
            status__in=(ReworkRequest.STATUS_PENDING, ReworkRequest.STATUS_IN_PROGRESS),
        ).exists():
            return _err('REWORK_ACTIVE', 'По этому браку уже есть активная переделка', http_status=422)
        qty_raw = request.data.get('quantity')
        qty_kg_raw = request.data.get('quantity_kg')
        year = tz.now().date().year
        last = ReworkRequest.objects.filter(rework_number__startswith=f'RWK-{year}-').order_by('-rework_number').first()
        try:
            last_n = int(last.rework_number.split('-')[-1]) if last else 0
        except (ValueError, IndexError):
            last_n = 0
        rw_n = f'RWK-{year}-{last_n + 1:04d}'
        user = request.user if request.user.is_authenticated else None

        if qty_kg_raw not in (None, '') and qty_raw in (None, ''):
            try:
                qkg = Decimal(str(qty_kg_raw))
            except Exception:
                return _err('INVALID_DECIMAL', 'Некорректное quantity_kg', http_status=422)
            try:
                with transaction.atomic():
                    rec = DefectRecord.objects.select_for_update().get(pk=record.pk)
                    qpcs, qkg_out = _reserve_defect_from_kg_only(rec, qkg)
                    rec.save(
                        update_fields=[
                            'sent_to_rework_quantity_pcs', 'quantity_pcs', 'quantity_kg',
                            'status', 'updated_at',
                        ],
                    )
                    rw = ReworkRequest.objects.create(
                        return_doc=None,
                        defect_record=rec,
                        original_sale=None,
                        product=rec.product,
                        quantity_pcs=qpcs,
                        quantity_kg=qkg_out,
                        status=ReworkRequest.STATUS_PENDING,
                        rework_number=rw_n,
                        created_by=user,
                    )
            except ValueError as e:
                return _err('INVALID_QUANTITY', str(e), http_status=422)
            record = rec
        else:
            rem_before = Decimal(str(record.quantity_pcs or 0))
            if rem_before <= 0:
                return _err('NO_QUANTITY', 'Нет остатка для отправки в переделку', http_status=422)
            try:
                send_qty = rem_before if qty_raw in (None, '') else Decimal(str(qty_raw))
            except Exception:
                return _err('INVALID_DECIMAL', 'Некорректное quantity', http_status=422)
            if send_qty <= 0:
                return _err('INVALID_QUANTITY', 'quantity должно быть > 0', http_status=422)
            if send_qty > rem_before + Decimal('0.0001'):
                return _err('QTY_TOO_HIGH', f'Нельзя отправить больше остатка ({rem_before})', http_status=422)
            eps = Decimal('0.0001')
            if send_qty >= rem_before - eps:
                try:
                    validate_defect_transition(record.status, DefectRecord.STATUS_SENT_TO_REWORK)
                except ValueError as e:
                    return _err('INVALID_STATUS', str(e), http_status=422)
            try:
                qmap = rework_quantities_from_defect_record(record, pcs_to_send=send_qty)
            except ValueError as e:
                return _err('NO_QUANTITY', str(e), http_status=422)
            kg_before = record.quantity_kg
            kg_before_d = Decimal(str(kg_before)) if kg_before is not None else None
            with transaction.atomic():
                record = DefectRecord.objects.select_for_update().get(pk=record.pk)
                record.sent_to_rework_quantity_pcs = Decimal(str(record.sent_to_rework_quantity_pcs or 0)) + send_qty
                record.recompute_remaining_pcs()
                if kg_before_d is not None and rem_before > 0 and kg_before_d > 0:
                    kg_delta = (send_qty / rem_before * kg_before_d).quantize(Decimal('0.0001'))
                    record.quantity_kg = max(Decimal('0'), (kg_before_d - kg_delta).quantize(Decimal('0.0001')))
                record.apply_terminal_status_from_counters()
                record.save(
                    update_fields=[
                        'sent_to_rework_quantity_pcs', 'quantity_pcs', 'quantity_kg',
                        'status', 'updated_at',
                    ],
                )
                rw = ReworkRequest.objects.create(
                    return_doc=None,
                    defect_record=record,
                    original_sale=None,
                    product=record.product,
                    quantity_pcs=qmap['quantity_pcs'],
                    quantity_kg=qmap['quantity_kg'],
                    status=ReworkRequest.STATUS_PENDING,
                    rework_number=rw_n,
                    created_by=user,
                )
        from .commercial_audit import log_commercial_audit
        log_commercial_audit(
            user=request.user,
            request=request,
            section='defects',
            description=f'Брак → переделка: defect #{record.pk}, {rw_n}',
            model_cls=DefectRecord,
            instance=record,
            payload_extra={'action': 'send_defect_to_rework', 'defect_id': record.pk, 'rework_id': rw.pk},
        )
        return Response({
            'defect': DefectRecordSerializer(record).data,
            'rework_request': ReworkRequestSerializer(rw).data,
        })

    @action(detail=True, methods=['post'], url_path='complete-rework')
    def complete_rework(self, request, pk=None):
        """Отключено: завершение переделки и выход на склад только через POST /api/rework-requests/{id}/complete/."""
        return Response(
            {
                'code': 'USE_REWORK_COMPLETE',
                'error': 'Завершение переделки только через POST /api/rework-requests/{id}/complete/',
                'detail': 'Завершение переделки только через POST /api/rework-requests/{id}/complete/',
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    @action(detail=True, methods=['post'], url_path='writeoff')
    def writeoff(self, request, pk=None):
        """Списать брак (полностью или часть — body quantity)."""
        from .state_machine import validate_defect_transition, validate_defect_writeoff_qty

        record = self.get_object()
        reason = request.data.get('writeoff_reason', '').strip()
        if not reason:
            return _err('MISSING_REASON', 'Укажите writeoff_reason — причина списания обязательна')
        try:
            validate_defect_writeoff_qty(record)
        except ValueError as e:
            return _err('INVALID_STATUS', str(e), http_status=422)
        rem = Decimal(str(record.quantity_pcs or 0))
        qty_raw = request.data.get('quantity')
        try:
            wo_qty = rem if qty_raw in (None, '') else Decimal(str(qty_raw))
        except Exception:
            return _err('INVALID_DECIMAL', 'Некорректное quantity', http_status=422)
        if wo_qty <= 0:
            return _err('INVALID_QUANTITY', 'quantity должно быть > 0', http_status=422)
        if wo_qty > rem + Decimal('0.0001'):
            return _err('QTY_TOO_HIGH', f'Нельзя списать больше остатка ({rem})', http_status=422)
        eps = Decimal('0.0001')
        if wo_qty >= rem - eps:
            try:
                validate_defect_transition(record.status, DefectRecord.STATUS_WRITTEN_OFF)
            except ValueError as e:
                return _err('INVALID_STATUS', str(e), http_status=422)
        kg_before_d = Decimal(str(record.quantity_kg)) if record.quantity_kg is not None else None
        try:
            with transaction.atomic():
                record = DefectRecord.objects.select_for_update().get(pk=record.pk)
                record.written_off_quantity_pcs = Decimal(str(record.written_off_quantity_pcs or 0)) + wo_qty
                record.recompute_remaining_pcs()
                if kg_before_d is not None and rem > 0 and kg_before_d > 0:
                    kg_delta = (wo_qty / rem * kg_before_d).quantize(Decimal('0.0001'))
                    record.quantity_kg = max(Decimal('0'), (kg_before_d - kg_delta).quantize(Decimal('0.0001')))
                record.apply_terminal_status_from_counters()
                record.writeoff_reason = reason
                record.save(
                    update_fields=[
                        'written_off_quantity_pcs', 'quantity_pcs', 'quantity_kg',
                        'status', 'writeoff_reason', 'updated_at',
                    ],
                )
                if record.warehouse_batch_id:
                    from apps.warehouse.models import WarehouseBatch
                    from apps.warehouse.stock_ops import apply_sale_to_warehouse_batch

                    wb = WarehouseBatch.objects.select_for_update().get(pk=record.warehouse_batch_id)
                    apply_sale_to_warehouse_batch(wb.pk, wo_qty, wb.inventory_form, None)
        except ValidationError as e:
            det = getattr(e, 'detail', None)
            if isinstance(det, dict):
                msg = '; '.join(
                    f'{k}: {(v[0] if isinstance(v, list) else v)}' for k, v in det.items()
                )
            else:
                msg = str(det or e)
            return _err('WAREHOUSE_APPLY', msg, http_status=422)
        from .commercial_audit import log_commercial_audit
        log_commercial_audit(
            user=request.user,
            request=request,
            section='defects',
            description=f'Списание брака #{record.pk}',
            model_cls=DefectRecord,
            instance=record,
            payload_extra={'action': 'writeoff_defect', 'defect_id': record.pk},
        )
        return Response(DefectRecordSerializer(record).data)

    @action(detail=True, methods=['post'], url_path='sell')
    def sell_defect(self, request, pk=None):
        """Продажа брака — создаёт Sale с is_defect_sale=True."""
        from django.utils import timezone
        from .state_machine import validate_defect_sell

        record = self.get_object()
        if record.status in (
            DefectRecord.STATUS_SOLD,
            DefectRecord.STATUS_WRITTEN_OFF,
            DefectRecord.STATUS_CLOSED,
        ):
            return _err('DEFECT_NOT_AVAILABLE', 'Нельзя продать запись брака в финальном статусе.', http_status=422)
        try:
            validate_defect_sell(record)
        except ValueError as e:
            return _err('INVALID_STATUS', str(e), http_status=422)
        client_id = request.data.get('client_id')
        price = request.data.get('price')
        quantity_raw = request.data.get('quantity')
        if client_id is None or client_id == '':
            return _err('MISSING_CLIENT', 'Укажите client_id', http_status=422)
        if price is None or str(price).strip() == '':
            return _err('MISSING_PRICE', 'Укажите price', http_status=422)
        if quantity_raw is None or str(quantity_raw).strip() == '':
            return _err('MISSING_QUANTITY', 'Укажите quantity (> 0, не больше остатка по браку)', http_status=422)
        try:
            client_pk = int(client_id)
        except Exception:
            return _err('MISSING_CLIENT', 'Укажите корректный client_id', http_status=422)
        client = Client.objects.filter(pk=client_pk).first()
        if client is None:
            return _err('MISSING_CLIENT', 'Клиент не найден', http_status=422)
        if not client.is_active:
            return _err('INACTIVE_CLIENT', 'Нельзя продавать брак неактивному клиенту.', http_status=422)
        try:
            price_d = Decimal(str(price))
        except Exception:
            return _err('INVALID_PRICE', 'Некорректный price', http_status=422)
        try:
            qty_d = Decimal(str(quantity_raw))
        except Exception:
            return _err('INVALID_QUANTITY', 'Некорректный quantity', http_status=422)
        if price_d <= 0:
            return _err('INVALID_PRICE', 'price должен быть > 0', http_status=422)
        if qty_d <= 0:
            return _err('INVALID_QUANTITY', 'quantity должен быть > 0', http_status=422)
        avail = Decimal(str(record.quantity_pcs or 0))
        if avail <= 0:
            return _err('DEFECT_NOT_AVAILABLE', 'Нет доступного остатка брака для продажи.', http_status=422)
        if qty_d > avail + Decimal('0.0001'):
            return _err('QUANTITY_EXCEEDED', f'Нельзя продать больше остатка по браку ({avail})', http_status=422)
        quantity = qty_d
        kg_before_d = Decimal(str(record.quantity_kg)) if record.quantity_kg is not None else None
        comment = request.data.get('comment', '')
        date = request.data.get('date') or timezone.now().date()

        wb_id = record.warehouse_batch_id
        stock_form = ''
        stock_quality = ''
        if wb_id:
            from apps.warehouse.models import WarehouseBatch

            try:
                wb_hdr = WarehouseBatch.objects.get(pk=wb_id)
            except WarehouseBatch.DoesNotExist:
                wb_hdr = None
                wb_id = None
            if wb_hdr is not None:
                stock_form = wb_hdr.inventory_form or ''
                stock_quality = (wb_hdr.quality or '').strip()

        from .sale_warehouse import apply_warehouse_for_sale

        try:
            with transaction.atomic():
                sale = Sale.objects.create(
                    order_number='',
                    product=record.product,
                    quantity=Decimal(str(quantity)),
                    sold_pieces=Decimal(str(quantity)),
                    price=price_d,
                    revenue=(price_d * quantity).quantize(Decimal('0.01')),
                    cost=Decimal('0'),
                    profit=Decimal('0'),
                    date=date,
                    comment=comment or f'Продажа брака #{record.id}',
                    is_defect_sale=True,
                    sale_status=Sale.STATUS_SHIPPED,
                    client_id=client.pk,
                    warehouse_batch_id=wb_id,
                    stock_form=stock_form,
                    stock_quality=stock_quality,
                )
                year = sale.date.year
                last = Sale.objects.filter(order_number__startswith=f'ORD-{year}-').exclude(pk=sale.pk).order_by('-order_number').first()
                try:
                    last_n = int(last.order_number.split('-')[-1]) if last else 0
                except (ValueError, IndexError):
                    last_n = 0
                sale.order_number = f'ORD-{year}-{last_n + 1:03d}'
                sale.save(update_fields=['order_number'])
                lt = (price_d * quantity).quantize(Decimal('0.01'))
                SaleLine.objects.create(
                    sale=sale,
                    product=record.product,
                    quantity=quantity,
                    unit_price=price_d,
                    line_total=lt,
                    cost=Decimal('0'),
                    profit=lt,
                )
                if wb_id:
                    applied = apply_warehouse_for_sale(Sale.objects.get(pk=sale.pk))
                    if not applied:
                        raise ValidationError({'warehouse_batch': 'Не удалось списать партию склада по продаже брака'})
                rec = DefectRecord.objects.select_for_update().get(pk=record.pk)
                rec.sold_quantity_pcs = Decimal(str(rec.sold_quantity_pcs or 0)) + quantity
                rec.recompute_remaining_pcs()
                if kg_before_d is not None and avail > 0 and kg_before_d > 0:
                    kg_delta = (quantity / avail * kg_before_d).quantize(Decimal('0.0001'))
                    rec.quantity_kg = max(Decimal('0'), (kg_before_d - kg_delta).quantize(Decimal('0.0001')))
                rec.apply_terminal_status_from_counters()
                rec.save(
                    update_fields=[
                        'sold_quantity_pcs', 'quantity_pcs', 'quantity_kg', 'status', 'updated_at',
                    ],
                )
        except ValidationError as e:
            det = getattr(e, 'detail', None)
            if isinstance(det, dict):
                msg = '; '.join(
                    f'{k}: {(v[0] if isinstance(v, list) else v)}' for k, v in det.items()
                )
            else:
                msg = str(det or e)
            return _err('WAREHOUSE_APPLY', msg, http_status=422)
        record.refresh_from_db()
        from .commercial_audit import log_commercial_audit
        log_commercial_audit(
            user=request.user,
            request=request,
            section='defects',
            description=f'Продажа брака #{record.pk} → sale #{sale.pk}',
            model_cls=DefectRecord,
            instance=record,
            payload_extra={'action': 'sell_defect', 'defect_id': record.pk, 'sale_id': sale.pk},
        )
        return Response({'sale_id': sale.id, 'sale_order_number': sale.order_number}, status=status.HTTP_201_CREATED)


# ─────────────────────────────────────────────────────────────────────────────
# REWORK REQUEST (Переделка)
# ─────────────────────────────────────────────────────────────────────────────

class ReworkRequestViewSet(ActivityLoggingMixin, viewsets.ModelViewSet):
    queryset = ReworkRequest.objects.select_related(
        'return_doc', 'defect_record', 'defect_record__warehouse_batch', 'original_sale', 'result_warehouse_batch', 'created_by',
    ).all()
    serializer_class = ReworkRequestSerializer
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'defects'
    activity_section = 'Переделки'
    activity_label = 'переделка'
    ordering_fields = ['id', 'created_at', 'status']
    filterset_fields = ['status', 'original_sale']

    def get_queryset(self):
        qs = super().get_queryset()
        if getattr(self, 'action', None) == 'list':
            return qs.filter(status=ReworkRequest.STATUS_COMPLETED)
        return qs

    def destroy(self, request, *args, **kwargs):
        return Response(
            {'code': 'DELETE_DISABLED', 'error': 'Удаление переделок отключено. Используйте cancel/complete.'},
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        serializer.save(created_by=user)
        inst = serializer.instance
        from apps.realtime.broadcast import schedule_push
        from django.db import transaction
        transaction.on_commit(lambda: schedule_push(
            resource='rework_request',
            action='created',
            entity_id=inst.pk,
            extra={'status': inst.status},
        ))

    def retrieve(self, request, *args, **kwargs):
        rework = self.get_object()
        data = ReworkRequestSerializer(rework, context={'request': request}).data
        from .state_machine import REWORK_TRANSITIONS
        data['linked_entities'] = {
            'return_doc': (
                {
                    'id': rework.return_doc_id,
                    'label': rework.return_doc.return_number or str(rework.return_doc_id),
                } if rework.return_doc_id else None
            ),
            'defect_record': (
                {
                    'id': rework.defect_record_id,
                    'label': f"#{rework.defect_record_id} {rework.defect_record.product}",
                } if rework.defect_record_id else None
            ),
            'original_sale': (
                {
                    'id': rework.original_sale_id,
                    'label': rework.original_sale.order_number,
                } if rework.original_sale_id else None
            ),
            'result_warehouse_batch': (
                {
                    'id': rework.result_warehouse_batch_id,
                    'label': f"#{rework.result_warehouse_batch_id} {rework.result_warehouse_batch.product}",
                } if rework.result_warehouse_batch_id else None
            ),
        }
        data['available_status_transitions'] = REWORK_TRANSITIONS.get(rework.status, [])
        data['available_actions'] = {
            'start': ReworkRequest.STATUS_IN_PROGRESS in REWORK_TRANSITIONS.get(rework.status, []),
            'complete': ReworkRequest.STATUS_COMPLETED in REWORK_TRANSITIONS.get(rework.status, []),
            'cancel': ReworkRequest.STATUS_CANCELED in REWORK_TRANSITIONS.get(rework.status, []),
        }
        return Response(data)

    @action(detail=False, methods=['get'], url_path='select-sources')
    def select_sources(self, request):
        from apps.warehouse.models import WarehouseBatch

        defects_qs = (
            DefectRecord.objects.exclude(
                status__in=(
                    DefectRecord.STATUS_SOLD,
                    DefectRecord.STATUS_WRITTEN_OFF,
                    DefectRecord.STATUS_CLOSED,
                ),
            )
            .select_related('warehouse_batch', 'profile')
            .order_by('-created_at', '-id')[:300]
        )
        defect_records = []
        for d in defects_qs:
            pcs = Decimal(str(d.quantity_pcs or 0))
            kg_raw = d.quantity_kg
            kg = Decimal(str(kg_raw)) if kg_raw is not None else Decimal('0')
            if pcs <= 0 and kg <= 0:
                continue
            if pcs > 0:
                qty_part = f'{api_decimal_str(pcs)} шт'
            elif kg > 0:
                qty_part = f'{api_decimal_str(kg)} кг'
            else:
                qty_part = '—'
            reason = (d.defect_reason or '').strip() or '—'
            reason_short = (reason[:80] + '...') if len(reason) > 80 else reason
            label = f'{d.product or "—"} — {qty_part} — {reason_short} — {d.get_source_type_display()}'
            product_s = (d.product or '').strip()
            profile_nm = (d.profile.name if d.profile_id else '') or ''
            defect_records.append({
                'id': d.id,
                'label': label,
                'product': product_s,
                'product_name': product_s,
                'profile_name': profile_nm,
                'quantity_pcs': api_decimal_str(pcs) if pcs > 0 else None,
                'quantity_kg': api_decimal_str(kg) if kg > 0 else None,
                'defect_reason': (d.defect_reason or '').strip(),
                'source_type': d.source_type,
                'source_label': defect_record_source_label(d),
                'display_quantity': api_decimal_str(pcs) if pcs > 0 else api_decimal_str(kg) if kg > 0 else None,
                'display_quantity_label': qty_part if qty_part != '—' else None,
                'available_quantity_pcs': api_decimal_str(pcs) if pcs > 0 else None,
                'status': d.status,
            })
        sales = list(
            Sale.objects.order_by('-date', '-id').values('id', 'order_number', 'product')[:300]
        )
        returns = list(
            Return.objects.order_by('-date', '-id').values('id', 'return_number')[:300]
        )
        result_batches = list(
            WarehouseBatch.objects.filter(
                status=WarehouseBatch.STATUS_AVAILABLE,
                quality=WarehouseBatch.QUALITY_GOOD,
                stock_bucket=WarehouseBatch.STOCK_BUCKET_STANDARD,
            )
            .order_by('-date', '-id')
            .values('id', 'product')[:300]
        )
        items = {
            'defect_records': defect_records,
            'original_sales': [
                {'id': s['id'], 'label': f"{s['order_number']} / {s['product']}"}
                for s in sales
            ],
            'returns': [
                {'id': r['id'], 'label': r['return_number'] or f"RET#{r['id']}"}
                for r in returns
            ],
            'result_warehouse_batches': [
                {'id': b['id'], 'label': f"#{b['id']} {b['product']}"}
                for b in result_batches
            ],
        }
        return Response({'items': items, **items})

    @action(detail=True, methods=['post'], url_path='start')
    def start_rework(self, request, pk=None):
        """Перевести переделку в статус «В работе»."""
        from .state_machine import validate_rework_transition
        rework = self.get_object()
        try:
            validate_rework_transition(rework.status, ReworkRequest.STATUS_IN_PROGRESS)
        except ValueError as e:
            return _err('INVALID_STATUS', str(e), http_status=422)
        rework.status = ReworkRequest.STATUS_IN_PROGRESS
        rework.save(update_fields=['status', 'updated_at'])
        return Response(ReworkRequestSerializer(rework).data)

    @action(detail=True, methods=['post'], url_path='complete')
    def complete(self, request, pk=None):
        """
        Завершить переделку: создать новую партию ГП (good или defect) по output_quantity / loss_quantity.
        result_warehouse_batch_id опционален (legacy); при отсутствии — партия создаётся в backend.
        """
        from django.utils import timezone as tz

        from apps.warehouse.models import WarehouseBatch
        from .state_machine import validate_rework_complete
        from django.db import transaction as db_transaction
        from apps.realtime.broadcast import schedule_push

        rework = self.get_object()
        if rework.status == ReworkRequest.STATUS_COMPLETED:
            return _err('REWORK_ALREADY_COMPLETED', 'Переделка уже завершена.', http_status=422)
        if rework.status == ReworkRequest.STATUS_CANCELED:
            return _err('REWORK_ALREADY_CANCELED', 'Отмененную переделку нельзя завершить.', http_status=422)
        if rework.status not in (ReworkRequest.STATUS_PENDING, ReworkRequest.STATUS_IN_PROGRESS):
            return _err('REWORK_COMPLETE_FORBIDDEN', 'Завершить переделку можно из статусов pending или in_progress.', http_status=422)
        try:
            validate_rework_complete(rework)
        except ValueError as e:
            return _err('INVALID_TRANSITION', str(e), http_status=422)

        out_raw = request.data.get('output_quantity', request.data.get('output_quantity_kg'))
        loss_raw = request.data.get('loss_quantity', request.data.get('loss_kg'))
        quality_raw = request.data.get('quality')
        if quality_raw is None or str(quality_raw).strip() == '':
            return _err('MISSING_FIELDS', 'Укажите quality.', http_status=422)
        out_quality = str(quality_raw).strip()
        if out_raw is None or loss_raw is None:
            return _err('MISSING_FIELDS', 'Укажите output_quantity (или output_quantity_kg) и loss_quantity (или loss_kg)')
        try:
            output_kg = Decimal(str(out_raw))
            loss_kg = Decimal(str(loss_raw))
        except Exception:
            return _err('INVALID_QUANTITY', 'Некорректные output_quantity/loss_quantity', http_status=422)
        if output_kg < 0 or loss_kg < 0:
            return _err('NEGATIVE_QUANTITY', 'output_quantity и loss_quantity должны быть >= 0', http_status=422)
        if out_quality not in (WarehouseBatch.QUALITY_GOOD, WarehouseBatch.QUALITY_DEFECT):
            return _err('INVALID_QUALITY', 'quality: good | defect', http_status=422)
        defect = rework.defect_record
        if defect is None:
            return _err('NO_DEFECT', 'Нет defect_record', http_status=422)

        eps = Decimal('0.0001')
        rw_kg = Decimal(str(rework.quantity_kg or 0)) if rework.quantity_kg is not None else Decimal('0')
        rw_pcs = Decimal(str(rework.quantity_pcs or 0)) if rework.quantity_pcs is not None else Decimal('0')
        coeff_raw = defect.kg_coefficient
        coeff = Decimal(str(coeff_raw)) if coeff_raw is not None else Decimal('0')

        input_kg = Decimal('0')
        if rw_kg > 0:
            input_kg = rw_kg
        elif rw_pcs > 0 and coeff > 0:
            input_kg = (rw_pcs * coeff).quantize(Decimal('0.0001'))
        elif rw_pcs > 0:
            input_kg = rw_pcs
        else:
            return _err('NO_INPUT', 'Не удалось определить массу входа переделки.', http_status=422)

        if input_kg > 0 and output_kg + loss_kg > input_kg + eps:
            return _err(
                'QTY_BOUNDS',
                f'output+loss не должно превышать вход ({api_decimal_str(input_kg)} кг)',
                http_status=422,
            )

        def _kg_to_wh_quantity(kg: Decimal) -> Decimal:
            if coeff > 0:
                return (kg / coeff).quantize(Decimal('0.0001'))
            return kg.quantize(Decimal('0.0001'))

        wb_quantity = _kg_to_wh_quantity(output_kg)

        tpl = getattr(defect, 'warehouse_batch', None)
        out_bucket = (
            WarehouseBatch.STOCK_BUCKET_REWORKED
            if out_quality == WarehouseBatch.QUALITY_GOOD
            else WarehouseBatch.STOCK_BUCKET_STANDARD
        )
        with transaction.atomic():
            wb = WarehouseBatch.objects.create(
                profile_id=defect.profile_id,
                product=defect.product,
                length_per_piece=tpl.length_per_piece if tpl else None,
                quantity=wb_quantity,
                cost_per_piece=tpl.cost_per_piece if tpl else Decimal('0'),
                cost_per_meter=tpl.cost_per_meter if tpl else Decimal('0'),
                date=tz.now().date(),
                source_batch_id=tpl.source_batch_id if tpl else None,
                inventory_form=tpl.inventory_form if tpl else WarehouseBatch.INVENTORY_UNPACKED,
                stock_bucket=out_bucket,
                quality=out_quality,
                defect_reason=(defect.defect_reason or '')[:2000] if out_quality == WarehouseBatch.QUALITY_DEFECT else '',
            )
            rework.output_quantity_kg = output_kg
            rework.loss_kg = loss_kg
            if input_kg > 0:
                rework.conversion_rate = (output_kg / input_kg).quantize(Decimal('0.000001'))
            rework.status = ReworkRequest.STATUS_COMPLETED
            rework.result_warehouse_batch = wb
            rework.save()
            if rework.defect_record_id:
                DefectRecord.objects.filter(pk=rework.defect_record_id).update(
                    status=DefectRecord.STATUS_REWORKED,
                )

        db_transaction.on_commit(lambda: schedule_push(
            resource='rework_request',
            action='updated',
            entity_id=rework.pk,
            extra={'status': ReworkRequest.STATUS_COMPLETED},
        ))
        from .commercial_audit import log_commercial_audit
        log_commercial_audit(
            user=request.user,
            request=request,
            section='rework',
            description=f'Завершение переделки {rework.rework_number}',
            model_cls=ReworkRequest,
            instance=rework,
            payload_extra={'action': 'complete_rework', 'rework_id': rework.pk},
        )
        rework.refresh_from_db()
        return Response(ReworkRequestSerializer(rework).data)

    @action(detail=True, methods=['post'], url_path='cancel')
    def cancel_rework(self, request, pk=None):
        """Отменить переделку."""
        from .state_machine import validate_rework_transition
        rework = self.get_object()
        if rework.status == ReworkRequest.STATUS_COMPLETED:
            return _err('REWORK_ALREADY_COMPLETED', 'Завершенную переделку нельзя отменить.', http_status=422)
        if rework.status == ReworkRequest.STATUS_CANCELED:
            return _err('REWORK_ALREADY_CANCELED', 'Переделка уже отменена.', http_status=422)
        if rework.status not in (ReworkRequest.STATUS_PENDING, ReworkRequest.STATUS_IN_PROGRESS):
            return _err('REWORK_CANCEL_FORBIDDEN', 'Отмена переделки разрешена только из pending/in_progress.', http_status=422)
        try:
            validate_rework_transition(rework.status, ReworkRequest.STATUS_CANCELED)
        except ValueError as e:
            return _err('INVALID_TRANSITION', str(e), http_status=422)
        rpc = Decimal(str(rework.quantity_pcs or 0))
        rkg = Decimal(str(rework.quantity_kg or 0)) if rework.quantity_kg is not None else Decimal('0')
        try:
            with transaction.atomic():
                rw2 = ReworkRequest.objects.select_for_update().get(pk=rework.pk)
                rw2.status = ReworkRequest.STATUS_CANCELED
                rw2.save(update_fields=['status', 'updated_at'])
                if rw2.defect_record_id and (rpc > 0 or rkg > 0):
                    dr = DefectRecord.objects.select_for_update().get(pk=rw2.defect_record_id)
                    dr.sent_to_rework_quantity_pcs = max(
                        Decimal('0'),
                        Decimal(str(dr.sent_to_rework_quantity_pcs or 0)) - rpc,
                    )
                    dr.recompute_remaining_pcs()
                    if dr.quantity_kg is not None and rkg > 0:
                        dr.quantity_kg = (Decimal(str(dr.quantity_kg or 0)) + rkg).quantize(Decimal('0.0001'))
                    if dr.status == DefectRecord.STATUS_SENT_TO_REWORK:
                        dr.status = DefectRecord.STATUS_ON_STOCK
                    dr.save(
                        update_fields=[
                            'sent_to_rework_quantity_pcs', 'quantity_pcs', 'quantity_kg',
                            'status', 'updated_at',
                        ],
                    )
        except Exception as e:
            return _err('WAREHOUSE_ROLLBACK', str(e), http_status=422)
        rework.refresh_from_db()
        return Response(ReworkRequestSerializer(rework).data)


# ─────────────────────────────────────────────────────────────────────────────
# PRICE LIST (Прайс-лист)
# ─────────────────────────────────────────────────────────────────────────────

class PriceListViewSet(ActivityLoggingMixin, viewsets.ModelViewSet):
    queryset = PriceList.objects.prefetch_related('product_prices', 'product_prices__profile').all()
    serializer_class = PriceListSerializer
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'sales'
    activity_section = 'Прайсы'
    activity_label = 'прайс-лист'

    @action(detail=False, methods=['get'], url_path='suggest-price')
    def suggest_price(self, request):
        """
        GET /api/price-lists/suggest-price/?client_id=&profile_id=&product=
        Возвращает рекомендованную цену по приоритету:
          1. Индивидуальная цена клиента
          2. Базовый прайс
          3. null
        """
        from .pricing import suggest_price, price_suggestion_to_dict
        from datetime import date

        client_id = request.query_params.get('client_id')
        profile_id = request.query_params.get('profile_id')
        product = request.query_params.get('product')
        date_raw = request.query_params.get('date')

        on_date = None
        if date_raw:
            try:
                from datetime import datetime
                on_date = datetime.strptime(date_raw[:10], '%Y-%m-%d').date()
            except (ValueError, TypeError):
                pass

        suggestion = suggest_price(
            client_id=int(client_id) if client_id else None,
            profile_id=int(profile_id) if profile_id else None,
            product=product,
            on_date=on_date,
        )
        return Response(price_suggestion_to_dict(suggestion))


# ─────────────────────────────────────────────────────────────────────────────
# CLIENT PRICE (Индивидуальные цены клиента)
# ─────────────────────────────────────────────────────────────────────────────

class ClientPriceViewSet(ActivityLoggingMixin, viewsets.ModelViewSet):
    queryset = ClientPrice.objects.select_related('client', 'profile').all()
    serializer_class = ClientPriceSerializer
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'sales'
    activity_section = 'Цены клиентов'
    activity_label = 'цена клиента'
    filterset_fields = ['client', 'profile']


# ─────────────────────────────────────────────────────────────────────────────
# ORDER RESERVATION (Резервы — отдельный список)
# ─────────────────────────────────────────────────────────────────────────────

class OrderReservationViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Только чтение. Управление резервами — через /orders/{id}/reserve/
    """
    queryset = OrderReservation.objects.select_related(
        'order_line', 'order_line__order', 'warehouse_batch', 'created_by',
    ).all()
    serializer_class = OrderReservationSerializer
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'client_orders'
    filterset_fields = ['status', 'order_line', 'warehouse_batch']


# ─────────────────────────────────────────────────────────────────────────────
# CLIENT FINANCIAL SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

class ClientFinancialSummaryView(viewsets.ViewSet):
    """
    Канонический endpoint:
    GET /api/client-financial-summary/?client_id=<id>

    Полная финансовая сводка по клиенту.
    """
    permission_classes = [IsAdminOrHasAccess]
    required_access_key = 'clients'

    @extend_schema(
        summary='Финансовая сводка клиента',
        description='Канонический endpoint: GET /api/client-financial-summary/?client_id=<id>',
        parameters=[
            OpenApiParameter(
                name='client_id',
                type=int,
                required=True,
                location=OpenApiParameter.QUERY,
                description='ID клиента',
            ),
        ],
    )
    def list(self, request):
        from config.api_numbers import api_decimal_str
        from .credit_check import check_credit_limit
        from .payment_status import payment_status
        from django.db.models import Sum
        from django.db.models.functions import Coalesce

        client_id = request.query_params.get('client_id')
        if not client_id:
            return _err('MISSING_PARAM', 'Укажите client_id')

        try:
            client = Client.objects.get(pk=client_id)
        except Client.DoesNotExist:
            return _err('NOT_FOUND', 'Клиент не найден', http_status=404)

        sales = Sale.objects.filter(client=client).exclude(
            sale_status__in=(Sale.STATUS_DRAFT, Sale.STATUS_CANCELED),
        )
        payments = Payment.objects.filter(client=client, status=Payment.STATUS_ACTIVE)

        total_revenue = sales.aggregate(t=Coalesce(Sum('revenue'), Decimal('0')))['t'] or Decimal('0')
        total_cost = (
            sales.exclude(is_defect_sale=True).aggregate(t=Coalesce(Sum('cost'), Decimal('0')))['t']
        ) or Decimal('0')
        total_profit = total_revenue - total_cost
        defect_revenue = (
            sales.filter(is_defect_sale=True).aggregate(t=Coalesce(Sum('revenue'), Decimal('0')))['t']
        ) or Decimal('0')

        total_incoming = (
            payments.filter(
                payment_type__in=[Payment.TYPE_PREPAYMENT, Payment.TYPE_PAYMENT, Payment.TYPE_SURCHARGE]
            ).aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']
        ) or Decimal('0')
        total_refunded = (
            payments.filter(payment_type=Payment.TYPE_REFUND)
            .aggregate(t=Coalesce(Sum('amount'), Decimal('0')))['t']
        ) or Decimal('0')
        net_paid = total_incoming - total_refunded

        client_debt = max(Decimal('0'), total_revenue - net_paid)
        client_advance = max(Decimal('0'), net_paid - total_revenue)

        credit_result = check_credit_limit(client)
        fin_payment_status = payment_status(
            total_due=total_revenue,
            net_paid=net_paid,
            total_incoming=total_incoming,
            total_refund=total_refunded,
        )

        return Response({
            'client_id': client.pk,
            'client_name': client.name,
            'payment_status': fin_payment_status,
            'total_revenue': api_decimal_str(total_revenue),
            'total_cost': api_decimal_str(total_cost),
            'total_profit': api_decimal_str(total_profit),
            'defect_revenue': api_decimal_str(defect_revenue),
            'total_paid_gross': api_decimal_str(total_incoming),
            'total_refunded': api_decimal_str(total_refunded),
            'total_paid_net': api_decimal_str(net_paid),
            'client_debt_money': api_decimal_str(client_debt),
            'client_advance_amount': api_decimal_str(client_advance),
            'credit_limit': api_decimal_str(credit_result.credit_limit) if credit_result.credit_limit is not None else None,
            'credit_limit_mode': credit_result.block_mode,
            'credit_available': api_decimal_str(credit_result.credit_available) if credit_result.credit_available is not None else None,
            'is_over_limit': credit_result.is_over_limit,
            'credit_warning': credit_result.warning,
        })
